#!/usr/bin/env python3
"""Evidence-first 2020 C credit-risk and bounded allocation pipeline."""

from __future__ import annotations

import argparse
import json
import math
from collections import defaultdict
from pathlib import Path
from typing import Any

import numpy as np
from openpyxl import load_workbook
from sklearn.ensemble import RandomForestClassifier
from sklearn.impute import SimpleImputer
from sklearn.linear_model import LogisticRegression
from sklearn.metrics import brier_score_loss, roc_auc_score
from sklearn.pipeline import Pipeline
from sklearn.preprocessing import StandardScaler

ATTACHMENT_1 = "raw/case_files/附件1：123家有信贷记录企业的相关数据.xlsx"
ATTACHMENT_2 = "raw/case_files/附件2：302家无信贷记录企业的相关数据.xlsx"
ATTACHMENT_3 = "raw/case_files/附件3：银行贷款年利率与客户流失率关系的统计数据.xlsx"
PLAN_PATH = "experiments/experiment_plan.json"

BASELINE = "BASELINE_ROBUST_SCORE_ALLOCATOR"
RIDGE = "RIDGE_LOGIT_EXPECTED_VALUE_ALLOCATOR"
FOREST = "RANDOM_FOREST_SCENARIO_ALLOCATOR"
CANDIDATES = (BASELINE, RIDGE, FOREST)

LOAN_MIN = 100_000
LOAN_MAX = 1_000_000
KNOWN_BUDGET = 60_000_000
UNKNOWN_BUDGET = 100_000_000
LGD_BASE = 0.60

FEATURE_NAMES = (
    "log_input_positive",
    "log_output_positive",
    "net_margin_ratio",
    "input_void_rate",
    "output_void_rate",
    "input_negative_rate",
    "output_negative_rate",
    "log_input_count",
    "log_output_count",
    "log_supplier_count",
    "log_customer_count",
    "input_partner_hhi",
    "output_partner_hhi",
    "input_month_cv",
    "output_month_cv",
    "log_active_months",
    "duplicate_rate",
    "output_to_input_log_ratio",
)

BASELINE_RISK_DIRECTIONS = np.asarray(
    [-1, -1, -1, 1, 1, 1, 1, -1, -1, -1, -1, 1, 1, 1, 1, -1, 1, -1],
    dtype=float,
)

SHOCK_DELTAS = {
    "mild": {
        "DIVERSIFIED": 0.010,
        "HIGH_VOLUME_LOW_MARGIN": 0.030,
        "INPUT_DEPENDENT": 0.025,
        "VOLATILE": 0.035,
    },
    "moderate": {
        "DIVERSIFIED": 0.025,
        "HIGH_VOLUME_LOW_MARGIN": 0.075,
        "INPUT_DEPENDENT": 0.065,
        "VOLATILE": 0.090,
    },
    "severe": {
        "DIVERSIFIED": 0.060,
        "HIGH_VOLUME_LOW_MARGIN": 0.150,
        "INPUT_DEPENDENT": 0.130,
        "VOLATILE": 0.180,
    },
}

SHOCK_CAPACITY = {
    "mild": {
        "DIVERSIFIED": 0.98,
        "HIGH_VOLUME_LOW_MARGIN": 0.92,
        "INPUT_DEPENDENT": 0.94,
        "VOLATILE": 0.90,
    },
    "moderate": {
        "DIVERSIFIED": 0.94,
        "HIGH_VOLUME_LOW_MARGIN": 0.80,
        "INPUT_DEPENDENT": 0.84,
        "VOLATILE": 0.76,
    },
    "severe": {
        "DIVERSIFIED": 0.85,
        "HIGH_VOLUME_LOW_MARGIN": 0.60,
        "INPUT_DEPENDENT": 0.66,
        "VOLATILE": 0.54,
    },
}


def parse_args() -> argparse.Namespace:
    parser = argparse.ArgumentParser()
    parser.add_argument("--case-root", type=Path, required=True)
    parser.add_argument("--candidate-id", choices=CANDIDATES, required=True)
    parser.add_argument("--seed", type=int, required=True)
    parser.add_argument("--output", type=Path, required=True)
    return parser.parse_args()


def json_load(path: Path) -> dict[str, Any]:
    value = json.loads(path.read_text(encoding="utf-8"))
    if not isinstance(value, dict):
        raise ValueError("JSON_OBJECT_REQUIRED")
    return value


def safe_float(value: Any) -> float:
    result = float(value)
    if not math.isfinite(result):
        raise ValueError("NONFINITE_NUMERIC_INPUT")
    return result


def _new_stats() -> dict[str, Any]:
    return {
        "raw_rows": 0,
        "duplicate_rows": 0,
        "valid_rows": 0,
        "void_rows": 0,
        "negative_rows": 0,
        "positive_total": 0.0,
        "negative_total_abs": 0.0,
        "net_total": 0.0,
        "partners": defaultdict(float),
        "months": defaultdict(float),
        "active_months": set(),
    }


def _normalized_invoice_key(row: tuple[Any, ...]) -> tuple[Any, ...]:
    values: list[Any] = []
    for value in row[:8]:
        if hasattr(value, "isoformat"):
            values.append(value.isoformat())
        elif isinstance(value, str):
            values.append(value.strip())
        else:
            values.append(value)
    return tuple(values)


def _aggregate_sheet(
    worksheet: Any,
    enterprise_ids: set[str],
    direction: str,
    stats: dict[str, dict[str, dict[str, Any]]],
) -> None:
    seen: set[tuple[Any, ...]] = set()
    for row in worksheet.iter_rows(min_row=2, values_only=True):
        if not any(value is not None for value in row):
            continue
        enterprise_id = str(row[0]).strip()
        if enterprise_id not in enterprise_ids:
            raise ValueError("UNREGISTERED_ENTERPRISE_IN_INVOICE")
        current = stats[enterprise_id][direction]
        current["raw_rows"] += 1
        key = _normalized_invoice_key(row)
        if key in seen:
            current["duplicate_rows"] += 1
            continue
        seen.add(key)
        status = str(row[7]).strip()
        if status == "作废发票":
            current["void_rows"] += 1
            continue
        if status != "有效发票":
            raise ValueError("UNKNOWN_INVOICE_STATUS")
        total = safe_float(row[6])
        current["valid_rows"] += 1
        if total < 0:
            current["negative_rows"] += 1
            current["negative_total_abs"] += abs(total)
        else:
            current["positive_total"] += total
        current["net_total"] += total
        partner = str(row[3]).strip()
        current["partners"][partner] += abs(total)
        date_value = row[2]
        if not hasattr(date_value, "year") or not hasattr(date_value, "month"):
            raise ValueError("INVOICE_DATE_INVALID")
        month = f"{date_value.year:04d}-{date_value.month:02d}"
        current["active_months"].add(month)
        current["months"][month] += max(total, 0.0)


def _coefficient_of_variation(values: list[float]) -> float:
    if not values:
        return 0.0
    array = np.asarray(values, dtype=float)
    mean = float(np.mean(array))
    if mean <= 0:
        return 0.0
    return float(np.std(array) / mean)


def _concentration(values: list[float]) -> float:
    total = sum(values)
    if total <= 0:
        return 1.0
    return float(sum((value / total) ** 2 for value in values))


def _feature_row(input_stats: dict[str, Any], output_stats: dict[str, Any]) -> list[float]:
    input_rows = max(int(input_stats["raw_rows"]), 1)
    output_rows = max(int(output_stats["raw_rows"]), 1)
    input_valid = max(int(input_stats["valid_rows"]), 1)
    output_valid = max(int(output_stats["valid_rows"]), 1)
    input_positive = float(input_stats["positive_total"])
    output_positive = float(output_stats["positive_total"])
    total_absolute = (
        input_positive
        + output_positive
        + float(input_stats["negative_total_abs"])
        + float(output_stats["negative_total_abs"])
        + 1.0
    )
    net_margin = (
        float(output_stats["net_total"]) - float(input_stats["net_total"])
    ) / total_absolute
    active_months = input_stats["active_months"] | output_stats["active_months"]
    duplicate_rate = (
        float(input_stats["duplicate_rows"]) + float(output_stats["duplicate_rows"])
    ) / (input_rows + output_rows)
    log_ratio = math.log1p(output_positive) - math.log1p(input_positive)
    return [
        math.log1p(input_positive),
        math.log1p(output_positive),
        float(np.clip(net_margin, -1.0, 1.0)),
        float(input_stats["void_rows"]) / input_rows,
        float(output_stats["void_rows"]) / output_rows,
        float(input_stats["negative_rows"]) / input_valid,
        float(output_stats["negative_rows"]) / output_valid,
        math.log1p(float(input_stats["valid_rows"])),
        math.log1p(float(output_stats["valid_rows"])),
        math.log1p(len(input_stats["partners"])),
        math.log1p(len(output_stats["partners"])),
        _concentration(list(input_stats["partners"].values())),
        _concentration(list(output_stats["partners"].values())),
        _coefficient_of_variation(list(input_stats["months"].values())),
        _coefficient_of_variation(list(output_stats["months"].values())),
        math.log1p(len(active_months)),
        duplicate_rate,
        float(np.clip(log_ratio, -20.0, 20.0)),
    ]


def load_enterprise_dataset(
    path: Path,
    label_ids: set[str],
) -> tuple[list[str], dict[str, str], dict[str, int], dict[str, list[float]], dict[str, int]]:
    workbook = load_workbook(path, read_only=True, data_only=True)
    info = workbook["企业信息"]
    enterprise_ids: list[str] = []
    ratings: dict[str, str] = {}
    labels: dict[str, int] = {}
    for row in info.iter_rows(min_row=2, values_only=True):
        if not any(value is not None for value in row):
            continue
        enterprise_id = str(row[0]).strip()
        enterprise_ids.append(enterprise_id)
        if len(row) >= 3 and row[2] is not None:
            ratings[enterprise_id] = str(row[2]).strip()
        if enterprise_id in label_ids:
            if len(row) < 4 or str(row[3]).strip() not in {"是", "否"}:
                raise ValueError("TRAIN_OR_VALIDATION_LABEL_MISSING")
            labels[enterprise_id] = int(str(row[3]).strip() == "是")
    if len(enterprise_ids) != len(set(enterprise_ids)):
        raise ValueError("DUPLICATE_ENTERPRISE_ID")
    if set(labels) != label_ids:
        raise ValueError("LABEL_SPLIT_ID_MISMATCH")
    stats = {
        enterprise_id: {"input": _new_stats(), "output": _new_stats()}
        for enterprise_id in enterprise_ids
    }
    _aggregate_sheet(workbook["进项发票信息"], set(enterprise_ids), "input", stats)
    _aggregate_sheet(workbook["销项发票信息"], set(enterprise_ids), "output", stats)
    features = {
        enterprise_id: _feature_row(stats[enterprise_id]["input"], stats[enterprise_id]["output"])
        for enterprise_id in enterprise_ids
    }
    duplicate_counts = {
        enterprise_id: int(
            stats[enterprise_id]["input"]["duplicate_rows"]
            + stats[enterprise_id]["output"]["duplicate_rows"]
        )
        for enterprise_id in enterprise_ids
    }
    workbook.close()
    return enterprise_ids, ratings, labels, features, duplicate_counts


def matrix(features: dict[str, list[float]], ids: list[str]) -> np.ndarray:
    result = np.asarray([features[enterprise_id] for enterprise_id in ids], dtype=float)
    if result.shape != (len(ids), len(FEATURE_NAMES)) or not np.isfinite(result).all():
        raise ValueError("FEATURE_MATRIX_INVALID")
    return result


def _sigmoid(values: np.ndarray) -> np.ndarray:
    clipped = np.clip(values, -30.0, 30.0)
    return 1.0 / (1.0 + np.exp(-clipped))


def fit_risk_model(
    candidate_id: str,
    seed: int,
    x_train: np.ndarray,
    y_train: np.ndarray,
    prediction_matrices: list[np.ndarray],
) -> tuple[list[np.ndarray], dict[str, Any]]:
    if len(set(y_train.tolist())) != 2:
        raise ValueError("TRAIN_LABEL_CLASS_MISSING")
    if candidate_id == BASELINE:
        medians = np.median(x_train, axis=0)
        q1 = np.quantile(x_train, 0.25, axis=0)
        q3 = np.quantile(x_train, 0.75, axis=0)
        scales = np.where(q3 - q1 > 1e-9, q3 - q1, 1.0)
        prior = float(np.clip(np.mean(y_train), 1e-4, 1 - 1e-4))
        intercept = math.log(prior / (1 - prior))

        def predict(values: np.ndarray) -> np.ndarray:
            robust_z = np.clip((values - medians) / scales, -5.0, 5.0)
            linear = intercept + 0.55 * (robust_z @ BASELINE_RISK_DIRECTIONS) / len(FEATURE_NAMES)
            return _sigmoid(linear)

        probabilities = [predict(values) for values in prediction_matrices]
        details = {
            "family": "fixed robust score",
            "fit_rows": int(len(y_train)),
            "train_default_rate": prior,
        }
    elif candidate_id == RIDGE:
        estimator = Pipeline(
            [
                ("imputer", SimpleImputer(strategy="median")),
                ("standardizer", StandardScaler()),
                (
                    "classifier",
                    LogisticRegression(
                        C=0.35,
                        max_iter=2000,
                        random_state=seed,
                        solver="liblinear",
                    ),
                ),
            ]
        )
        estimator.fit(x_train, y_train)
        probabilities = [estimator.predict_proba(values)[:, 1] for values in prediction_matrices]
        coefficients = estimator.named_steps["classifier"].coef_[0]
        details = {
            "family": "L2 logistic regression",
            "fit_rows": int(len(y_train)),
            "coefficient_l2_norm": float(np.linalg.norm(coefficients)),
        }
    elif candidate_id == FOREST:
        estimator = Pipeline(
            [
                ("imputer", SimpleImputer(strategy="median")),
                (
                    "classifier",
                    RandomForestClassifier(
                        n_estimators=320,
                        max_depth=6,
                        max_features="sqrt",
                        min_samples_leaf=4,
                        class_weight="balanced_subsample",
                        random_state=seed,
                        n_jobs=1,
                    ),
                ),
            ]
        )
        estimator.fit(x_train, y_train)
        probabilities = [estimator.predict_proba(values)[:, 1] for values in prediction_matrices]
        importances = estimator.named_steps["classifier"].feature_importances_
        top = sorted(zip(FEATURE_NAMES, importances, strict=True), key=lambda item: -item[1])[:5]
        details = {
            "family": "class-weighted random forest",
            "fit_rows": int(len(y_train)),
            "top_feature_importance": [
                {"feature": name, "importance": float(value)} for name, value in top
            ],
        }
    else:
        raise ValueError("UNKNOWN_CANDIDATE")
    clipped_probabilities = [np.clip(values, 0.005, 0.995) for values in probabilities]
    if not all(np.isfinite(values).all() for values in clipped_probabilities):
        raise ValueError("NONFINITE_RISK_PROBABILITY")
    return clipped_probabilities, details


def validation_metric_values(y_true: np.ndarray, probabilities: np.ndarray) -> dict[str, float]:
    brier = float(brier_score_loss(y_true, probabilities))
    auc = float(roc_auc_score(y_true, probabilities))
    composite = brier + 0.25 * (1.0 - auc)
    if not all(math.isfinite(value) for value in (brier, auc, composite)):
        raise ValueError("VALIDATION_METRIC_NONFINITE")
    return {
        "validation_brier_loss": round(brier, 10),
        "validation_composite_loss": round(composite, 10),
        "validation_roc_auc": round(auc, 10),
    }


def validation_perturbations(
    candidate_id: str,
    seed: int,
    x_train: np.ndarray,
    y_train: np.ndarray,
    x_validation: np.ndarray,
    y_validation: np.ndarray,
    base_probabilities: np.ndarray,
) -> list[dict[str, Any]]:
    duplicate_index = FEATURE_NAMES.index("duplicate_rate")
    neutral_train = x_train.copy()
    neutral_validation = x_validation.copy()
    duplicate_reference = float(np.median(x_train[:, duplicate_index]))
    neutral_train[:, duplicate_index] = duplicate_reference
    neutral_validation[:, duplicate_index] = duplicate_reference
    (neutral_probabilities,), _ = fit_risk_model(
        candidate_id,
        seed,
        neutral_train,
        y_train,
        [neutral_validation],
    )

    scaled_validation = x_validation.copy()
    for index in (0, 1):
        scaled_validation[:, index] = np.log1p(1.10 * np.expm1(scaled_validation[:, index]))
    (scaled_probabilities,), _ = fit_risk_model(
        candidate_id,
        seed,
        x_train,
        y_train,
        [scaled_validation],
    )

    prior = float(np.mean(y_train))
    shrunk_probabilities = np.clip(0.90 * base_probabilities + 0.10 * prior, 0.005, 0.995)
    recomputations = (
        ("DUPLICATE_SIGNAL_NEUTRALIZED", neutral_probabilities),
        ("VALIDATION_MONETARY_SCALE_PLUS_10_PERCENT", scaled_probabilities),
        ("PROBABILITY_SHRINK_10_PERCENT_TO_TRAIN_PRIOR", shrunk_probabilities),
    )
    return [
        {
            "evidence": "DETERMINISTIC_RECOMPUTATION_FROM_BOUND_INPUTS",
            "metric": "validation_composite_loss",
            "perturbation_id": perturbation_id,
            "result": validation_metric_values(y_validation, probabilities)[
                "validation_composite_loss"
            ],
        }
        for perturbation_id, probabilities in recomputations
    ]


def load_rate_table(path: Path) -> tuple[list[float], dict[str, list[float]], dict[str, int]]:
    workbook = load_workbook(path, read_only=True, data_only=True)
    worksheet = workbook["Sheet1"]
    rows = [
        row[:4] for row in worksheet.iter_rows(min_row=3, values_only=True) if row[0] is not None
    ]
    rates = [safe_float(row[0]) for row in rows]
    monotone: dict[str, list[float]] = {}
    violations: dict[str, int] = {}
    for index, grade in enumerate(("A", "B", "C"), start=1):
        raw = [safe_float(row[index]) for row in rows]
        violations[grade] = sum(
            raw[position] < raw[position - 1] for position in range(1, len(raw))
        )
        monotone[grade] = np.maximum.accumulate(np.asarray(raw, dtype=float)).tolist()
    workbook.close()
    if len(rates) != 29 or rates[0] != 0.04 or rates[-1] != 0.15:
        raise ValueError("RATE_GRID_INVALID")
    return rates, monotone, violations


def choose_rate(
    risk: float,
    grade: str,
    rates: list[float],
    attrition: dict[str, list[float]],
    lgd: float,
) -> tuple[float, float, float]:
    if grade not in attrition:
        raise ValueError("RATE_GRADE_INVALID")
    margins = [
        (1.0 - attrition[grade][index]) * ((1.0 - risk) * rate - risk * lgd)
        for index, rate in enumerate(rates)
    ]
    best_index = max(range(len(rates)), key=lambda index: (margins[index], -rates[index]))
    return rates[best_index], attrition[grade][best_index], margins[best_index]


def allocate_exact(ids: list[str], budget: int, weights: dict[str, float]) -> dict[str, int]:
    if len(ids) != len(set(ids)) or not ids:
        raise ValueError("ALLOCATION_RECIPIENT_SET_INVALID")
    if budget < len(ids) * LOAN_MIN or budget > len(ids) * LOAN_MAX:
        raise ValueError("ALLOCATION_BUDGET_INFEASIBLE")
    continuous = {enterprise_id: float(LOAN_MIN) for enterprise_id in ids}
    remaining = float(budget - len(ids) * LOAN_MIN)
    active = set(ids)
    while remaining > 1e-8 and active:
        denominator = sum(max(float(weights[enterprise_id]), 1e-12) for enterprise_id in active)
        proposed = {
            enterprise_id: remaining * max(float(weights[enterprise_id]), 1e-12) / denominator
            for enterprise_id in active
        }
        saturated = [
            enterprise_id
            for enterprise_id in active
            if proposed[enterprise_id] >= LOAN_MAX - continuous[enterprise_id] - 1e-9
        ]
        if saturated:
            for enterprise_id in saturated:
                addition = LOAN_MAX - continuous[enterprise_id]
                continuous[enterprise_id] += addition
                remaining -= addition
                active.remove(enterprise_id)
        else:
            for enterprise_id in active:
                continuous[enterprise_id] += proposed[enterprise_id]
            remaining = 0.0
    allocations = {
        enterprise_id: int(math.floor(value)) for enterprise_id, value in continuous.items()
    }
    residual = budget - sum(allocations.values())
    fractional_order = sorted(
        ids,
        key=lambda enterprise_id: (
            -(continuous[enterprise_id] - allocations[enterprise_id]),
            enterprise_id,
        ),
    )
    for enterprise_id in fractional_order:
        if residual == 0:
            break
        if allocations[enterprise_id] < LOAN_MAX:
            allocations[enterprise_id] += 1
            residual -= 1
    if residual != 0:
        raise ValueError("ALLOCATION_INTEGER_RECONCILIATION_FAILED")
    return allocations


def infer_grades(ids: list[str], risks: dict[str, float]) -> dict[str, str]:
    values = np.asarray([risks[enterprise_id] for enterprise_id in ids], dtype=float)
    lower, upper = np.quantile(values, [0.25, 0.65])
    return {
        enterprise_id: (
            "A" if risks[enterprise_id] <= lower else "B" if risks[enterprise_id] <= upper else "C"
        )
        for enterprise_id in ids
    }


def build_schedule(
    ids: list[str],
    risks: dict[str, float],
    grades: dict[str, str],
    features: dict[str, list[float]],
    rates: list[float],
    attrition: dict[str, list[float]],
    budget: int,
    *,
    lgd: float,
    known_ratings: bool,
    capacity_multipliers: dict[str, float] | None = None,
    categories: dict[str, str] | None = None,
) -> tuple[list[dict[str, Any]], dict[str, Any]]:
    economics: dict[str, tuple[float, float, float]] = {}
    for enterprise_id in ids:
        if known_ratings and grades[enterprise_id] == "D":
            continue
        economics[enterprise_id] = choose_rate(
            risks[enterprise_id], grades[enterprise_id], rates, attrition, lgd
        )
    if known_ratings:
        recipients = sorted(economics)
    else:
        ranked = sorted(
            economics,
            key=lambda enterprise_id: (
                -economics[enterprise_id][2],
                risks[enterprise_id],
                enterprise_id,
            ),
        )
        positive_low_risk = [
            enterprise_id
            for enterprise_id in ranked
            if economics[enterprise_id][2] > 0 and risks[enterprise_id] <= 0.45
        ]
        recipient_count = max(100, min(200, len(positive_low_risk)))
        recipients = ranked[:recipient_count]
    weights: dict[str, float] = {}
    for enterprise_id in recipients:
        revenue_scale = math.sqrt(max(math.expm1(features[enterprise_id][1]), 0.0) + 1.0)
        capacity = 1.0
        if capacity_multipliers is not None and categories is not None:
            capacity = capacity_multipliers[categories[enterprise_id]]
        weights[enterprise_id] = max(economics[enterprise_id][2], 1e-6) * revenue_scale * capacity
    allocations = allocate_exact(recipients, budget, weights)
    records: list[dict[str, Any]] = []
    objective = 0.0
    for enterprise_id in ids:
        amount = allocations.get(enterprise_id, 0)
        if amount:
            rate, loss_rate, margin = economics[enterprise_id]
            objective += amount * margin
            record = {
                "annual_interest_rate": round(rate, 6),
                "attrition_rate": round(loss_rate, 8),
                "enterprise_id": enterprise_id,
                "expected_net_margin_per_yuan": round(margin, 10),
                "grade_or_proxy": grades[enterprise_id],
                "loan_amount_yuan": amount,
                "loan_decision": "LEND",
                "loan_term_years": 1,
                "risk_probability": round(risks[enterprise_id], 10),
            }
        else:
            record = {
                "annual_interest_rate": None,
                "attrition_rate": None,
                "enterprise_id": enterprise_id,
                "expected_net_margin_per_yuan": None,
                "grade_or_proxy": grades[enterprise_id],
                "loan_amount_yuan": 0,
                "loan_decision": "NO_LOAN",
                "loan_term_years": 0,
                "risk_probability": round(risks[enterprise_id], 10),
            }
        if categories is not None:
            record["transaction_category_proxy"] = categories[enterprise_id]
        records.append(record)
    summary = audit_schedule(records, budget, known_ratings=known_ratings)
    summary.update(
        {
            "expected_net_return_yuan": round(objective, 2),
            "lgd_assumption": lgd,
            "recipient_count": len(recipients),
        }
    )
    return records, summary


def audit_schedule(
    records: list[dict[str, Any]], budget: int, *, known_ratings: bool
) -> dict[str, Any]:
    violations: list[str] = []
    total = sum(int(record["loan_amount_yuan"]) for record in records)
    if total != budget:
        violations.append("BUDGET_NOT_EXACT")
    ids = [str(record["enterprise_id"]) for record in records]
    if len(ids) != len(set(ids)):
        violations.append("DUPLICATE_ENTERPRISE_DECISION")
    for record in records:
        amount = int(record["loan_amount_yuan"])
        if record["loan_decision"] == "LEND":
            rate = record["annual_interest_rate"]
            if amount < LOAN_MIN or amount > LOAN_MAX:
                violations.append("LOAN_BOUND_VIOLATION")
            if not isinstance(rate, (int, float)) or not 0.04 <= float(rate) <= 0.15:
                violations.append("RATE_BOUND_VIOLATION")
            if record["loan_term_years"] != 1:
                violations.append("TERM_VIOLATION")
        elif amount != 0 or record["annual_interest_rate"] is not None:
            violations.append("NO_LOAN_RECORD_INVALID")
        if known_ratings and record["grade_or_proxy"] == "D" and amount != 0:
            violations.append("KNOWN_D_LOAN_PROHIBITED")
    if violations:
        raise ValueError(";".join(sorted(set(violations))))
    return {
        "budget_exact": True,
        "budget_yuan": budget,
        "loan_bounds_pass": True,
        "rate_bounds_pass": True,
        "term_pass": True,
        "known_D_rule_pass": True,
        "total_allocated_yuan": total,
        "violations": [],
    }


def transaction_categories(ids: list[str], features: dict[str, list[float]]) -> dict[str, str]:
    values = matrix(features, ids)
    revenue_median = float(np.median(values[:, 1]))
    margin_median = float(np.median(values[:, 2]))
    input_ratio_median = float(np.median(-values[:, 17]))
    volatility_median = float(np.median(values[:, 14]))
    categories: dict[str, str] = {}
    for enterprise_id in ids:
        row = features[enterprise_id]
        if row[14] > volatility_median:
            category = "VOLATILE"
        elif -row[17] > input_ratio_median:
            category = "INPUT_DEPENDENT"
        elif row[1] > revenue_median and row[2] < margin_median:
            category = "HIGH_VOLUME_LOW_MARGIN"
        else:
            category = "DIVERSIFIED"
        categories[enterprise_id] = category
    return categories


def rating_priors(
    train_ids: list[str], labels: dict[str, int], ratings: dict[str, str]
) -> dict[str, float]:
    global_rate = sum(labels[enterprise_id] for enterprise_id in train_ids) / len(train_ids)
    priors: dict[str, float] = {}
    for grade in ("A", "B", "C", "D"):
        grade_ids = [
            enterprise_id for enterprise_id in train_ids if ratings[enterprise_id] == grade
        ]
        defaults = sum(labels[enterprise_id] for enterprise_id in grade_ids)
        priors[grade] = (defaults + 2.0 * global_rate) / (len(grade_ids) + 2.0)
    return priors


def feature_shift(
    x_train: np.ndarray, x_unknown: np.ndarray
) -> tuple[dict[str, float], list[dict[str, Any]]]:
    train_mean = np.mean(x_train, axis=0)
    train_std = np.std(x_train, axis=0)
    unknown_mean = np.mean(x_unknown, axis=0)
    standardized = np.abs(unknown_mean - train_mean) / np.where(train_std > 1e-9, train_std, 1.0)
    order = np.argsort(-standardized)
    top = [
        {
            "feature": FEATURE_NAMES[index],
            "standardized_mean_difference": round(float(standardized[index]), 6),
        }
        for index in order[:5]
    ]
    return {
        "maximum_standardized_mean_difference": round(float(np.max(standardized)), 6),
        "median_standardized_mean_difference": round(float(np.median(standardized)), 6),
    }, top


def risk_deciles(records: list[dict[str, Any]]) -> list[dict[str, Any]]:
    ordered = sorted(
        records, key=lambda record: (record["risk_probability"], record["enterprise_id"])
    )
    groups = np.array_split(np.asarray(ordered, dtype=object), 10)
    result: list[dict[str, Any]] = []
    for index, group in enumerate(groups, start=1):
        items = list(group)
        result.append(
            {
                "decile": index,
                "enterprise_count": len(items),
                "loan_total_yuan": sum(int(item["loan_amount_yuan"]) for item in items),
                "mean_risk": round(float(np.mean([item["risk_probability"] for item in items])), 8),
            }
        )
    return result


def main() -> None:
    args = parse_args()
    root = args.case_root
    plan = json_load(root / PLAN_PATH)["content"]
    if args.candidate_id not in plan["candidate_ids"] or args.seed not in plan["random_seeds"]:
        raise ValueError("EXECUTION_NOT_PREREGISTERED")
    train_ids = list(plan["splits"]["train"])
    validation_ids = list(plan["splits"]["validation"])
    test_ids = list(plan["splits"]["test"])
    if (
        set(train_ids) & set(validation_ids)
        or set(train_ids) & set(test_ids)
        or set(validation_ids) & set(test_ids)
    ):
        raise ValueError("SPLIT_OVERLAP")
    label_ids = set(train_ids) | set(validation_ids)
    known_ids, known_ratings, known_labels, known_features, known_duplicates = (
        load_enterprise_dataset(root / ATTACHMENT_1, label_ids)
    )
    unknown_ids, _, _, unknown_features, unknown_duplicates = load_enterprise_dataset(
        root / ATTACHMENT_2, set()
    )
    if set(known_ids) != set(train_ids) | set(validation_ids) | set(test_ids):
        raise ValueError("KNOWN_SPLIT_COVERAGE_MISMATCH")
    x_train = matrix(known_features, train_ids)
    x_validation = matrix(known_features, validation_ids)
    x_known = matrix(known_features, known_ids)
    x_unknown = matrix(unknown_features, unknown_ids)
    y_train = np.asarray([known_labels[enterprise_id] for enterprise_id in train_ids], dtype=int)
    y_validation = np.asarray(
        [known_labels[enterprise_id] for enterprise_id in validation_ids], dtype=int
    )
    (p_validation, p_known_transfer, p_unknown), model_details = fit_risk_model(
        args.candidate_id,
        args.seed,
        x_train,
        y_train,
        [x_validation, x_known, x_unknown],
    )
    validation_metrics = validation_metric_values(y_validation, p_validation)
    robustness_perturbations = validation_perturbations(
        args.candidate_id,
        args.seed,
        x_train,
        y_train,
        x_validation,
        y_validation,
        p_validation,
    )

    priors = rating_priors(train_ids, known_labels, known_ratings)
    known_risks = {
        enterprise_id: float(
            np.clip(
                0.70 * p_known_transfer[index] + 0.30 * priors[known_ratings[enterprise_id]],
                0.005,
                0.995,
            )
        )
        for index, enterprise_id in enumerate(known_ids)
    }
    unknown_risks = {
        enterprise_id: float(p_unknown[index]) for index, enterprise_id in enumerate(unknown_ids)
    }
    rates, attrition, rate_violations = load_rate_table(root / ATTACHMENT_3)

    known_records, known_summary = build_schedule(
        known_ids,
        known_risks,
        known_ratings,
        known_features,
        rates,
        attrition,
        KNOWN_BUDGET,
        lgd=LGD_BASE,
        known_ratings=True,
    )
    known_frontier: list[dict[str, Any]] = []
    for budget in plan["decision_budgets_yuan"]["known_budget_frontier"]:
        _, summary = build_schedule(
            known_ids,
            known_risks,
            known_ratings,
            known_features,
            rates,
            attrition,
            int(budget),
            lgd=LGD_BASE,
            known_ratings=True,
        )
        known_frontier.append(summary)

    unknown_grades = infer_grades(unknown_ids, unknown_risks)
    unknown_records, unknown_summary = build_schedule(
        unknown_ids,
        unknown_risks,
        unknown_grades,
        unknown_features,
        rates,
        attrition,
        UNKNOWN_BUDGET,
        lgd=LGD_BASE,
        known_ratings=False,
    )

    categories = transaction_categories(unknown_ids, unknown_features)
    scenario_summaries: list[dict[str, Any]] = []
    moderate_records: list[dict[str, Any]] = []
    for scenario in ("mild", "moderate", "severe"):
        stressed_risks = {
            enterprise_id: float(
                np.clip(
                    unknown_risks[enterprise_id]
                    + SHOCK_DELTAS[scenario][categories[enterprise_id]],
                    0.005,
                    0.995,
                )
            )
            for enterprise_id in unknown_ids
        }
        stressed_grades = infer_grades(unknown_ids, stressed_risks)
        records, summary = build_schedule(
            unknown_ids,
            stressed_risks,
            stressed_grades,
            unknown_features,
            rates,
            attrition,
            UNKNOWN_BUDGET,
            lgd=LGD_BASE,
            known_ratings=False,
            capacity_multipliers=SHOCK_CAPACITY[scenario],
            categories=categories,
        )
        summary["scenario"] = scenario
        summary["mean_stressed_risk"] = round(float(np.mean(list(stressed_risks.values()))), 8)
        scenario_summaries.append(summary)
        if scenario == "moderate":
            moderate_records = records

    lgd_sensitivity: list[dict[str, Any]] = []
    for lgd in (0.40, 0.60, 0.80):
        _, summary = build_schedule(
            unknown_ids,
            unknown_risks,
            unknown_grades,
            unknown_features,
            rates,
            attrition,
            UNKNOWN_BUDGET,
            lgd=lgd,
            known_ratings=False,
        )
        lgd_sensitivity.append(summary)
    risk_scale_sensitivity: list[dict[str, Any]] = []
    for scale in (0.80, 1.00, 1.20):
        scaled = {
            enterprise_id: float(np.clip(risk * scale, 0.005, 0.995))
            for enterprise_id, risk in unknown_risks.items()
        }
        scaled_grades = infer_grades(unknown_ids, scaled)
        _, summary = build_schedule(
            unknown_ids,
            scaled,
            scaled_grades,
            unknown_features,
            rates,
            attrition,
            UNKNOWN_BUDGET,
            lgd=LGD_BASE,
            known_ratings=False,
        )
        summary["risk_multiplier"] = scale
        risk_scale_sensitivity.append(summary)

    shift_summary, shift_top = feature_shift(x_train, x_unknown)
    category_counts = {
        category: sum(value == category for value in categories.values())
        for category in sorted(set(categories.values()))
    }
    final_metrics = {
        "known_expected_net_return_yuan": known_summary["expected_net_return_yuan"],
        "known_representative_budget_yuan": KNOWN_BUDGET,
        "moderate_shock_expected_net_return_yuan": next(
            item["expected_net_return_yuan"]
            for item in scenario_summaries
            if item["scenario"] == "moderate"
        ),
        "unknown_expected_net_return_yuan": unknown_summary["expected_net_return_yuan"],
        "unknown_recipient_count": unknown_summary["recipient_count"],
        **validation_metrics,
    }
    if args.output.is_absolute() or ".." in args.output.parts:
        raise ValueError("OUTPUT_PATH_MUST_BE_CASE_RELATIVE")
    output_relative = args.output.as_posix()
    if len(args.output.parts) < 3 or args.output.parts[0] != "runs":
        raise ValueError("OUTPUT_PATH_MUST_BE_RUN_BOUND")
    claim_scope = (
        "在绑定附件与预注册假设范围内，模型提供企业信贷风险 proxy、约束可行的一年期贷款策略和"
        "交易结构类别的突发事件情景调整；不构成外部校准或真实行业识别。"
    )
    output = {
        "candidate_id": args.candidate_id,
        "claim_scope": claim_scope,
        "data_diagnostics": {
            "feature_count": len(FEATURE_NAMES),
            "feature_names": list(FEATURE_NAMES),
            "feature_shift": shift_summary,
            "feature_shift_top": shift_top,
            "known_enterprise_count": len(known_ids),
            "known_exact_duplicates_removed": sum(known_duplicates.values()),
            "rate_monotonicity_violations": rate_violations,
            "test_enterprise_count_unscored": len(test_ids),
            "test_labels_accessed": False,
            "train_enterprise_count": len(train_ids),
            "unknown_enterprise_count": len(unknown_ids),
            "unknown_exact_duplicates_removed": sum(unknown_duplicates.values()),
            "validation_enterprise_count": len(validation_ids),
        },
        "decision_sensitivity": {
            "allocation_feasibility": {
                "known_representative": {
                    key: value
                    for key, value in known_summary.items()
                    if key not in {"expected_net_return_yuan", "lgd_assumption", "recipient_count"}
                },
                "moderate_shock": {
                    key: value
                    for key, value in next(
                        item for item in scenario_summaries if item["scenario"] == "moderate"
                    ).items()
                    if key
                    not in {
                        "expected_net_return_yuan",
                        "lgd_assumption",
                        "mean_stressed_risk",
                        "recipient_count",
                        "scenario",
                    }
                },
                "unknown": {
                    key: value
                    for key, value in unknown_summary.items()
                    if key not in {"expected_net_return_yuan", "lgd_assumption", "recipient_count"}
                },
            },
            "lgd_sensitivity": lgd_sensitivity,
            "risk_scale_sensitivity": risk_scale_sensitivity,
            "shock_sensitivity": scenario_summaries,
        },
        "figure_ready_data": [
            {
                "data": risk_deciles(unknown_records),
                "figure_id": "FIG-UNKNOWN-RISK-DECILES",
                "purpose": "展示无信贷记录企业风险 proxy、获贷额与风险分位的关系",
            },
            {
                "data": known_frontier,
                "figure_id": "FIG-KNOWN-BUDGET-FRONTIER",
                "purpose": "展示问题1固定总额变化下的可行额度与期望净收益",
            },
            {
                "data": scenario_summaries,
                "figure_id": "FIG-SHOCK-SCENARIO-SUMMARY",
                "purpose": "比较轻度、中度、重度交易结构冲击的调整结果",
            },
        ],
        "final_metrics": final_metrics,
        "limitations": [
            "附件1仅 123 家且违约标签与 D 级评级高度混杂；非 D 级违约极少。",
            "附件2没有违约真值，不能检验域外校准或声称真实违约概率。",
            "题面未给问题1固定总额，6000万元是可替换的代表性预算。",
            "无真实行业、资产负债、抵押物、贷款回收和宏观冲击数据。",
            "流失率为2019年评级组聚合关系，单调包络不是个体因果模型。",
        ],
        "model_summary": model_details,
        "problem_outputs": {
            "known_123": {
                "budget_frontier": known_frontier,
                "representative_schedule": known_records,
                "summary": known_summary,
            },
            "shock_adjustment": {
                "category_counts": category_counts,
                "category_disclaimer": "transaction-structure proxy; not actual industry",
                "moderate_schedule": moderate_records,
                "scenario_summaries": scenario_summaries,
            },
            "unknown_302": {
                "schedule": unknown_records,
                "summary": unknown_summary,
            },
        },
        "random_seed": args.seed,
        "requirement_claims": {
            "REQ-2020C-01": {
                "claim_id": "CLAIM-2020C-01",
                "claim_text": claim_scope,
                "evidence_artifact_ids": [output_relative],
            },
            "REQ-2020C-02": {
                "claim_id": "CLAIM-2020C-02",
                "claim_text": "问题1策略对可行固定总额参数化，并给出6000万元代表性可行表。",
                "evidence_artifact_ids": [output_relative],
            },
            "REQ-2020C-03": {
                "claim_id": "CLAIM-2020C-03",
                "claim_text": "302家无记录企业均获得由同构发票特征迁移得到的风险proxy。",
                "evidence_artifact_ids": [output_relative],
            },
            "REQ-2020C-04": {
                "claim_id": "CLAIM-2020C-04",
                "claim_text": "问题2逐企业策略通过1亿元总额、额度、利率和期限独立可行性检查。",
                "evidence_artifact_ids": [output_relative],
            },
            "REQ-2020C-05": {
                "claim_id": "CLAIM-2020C-05",
                "claim_text": "问题3按交易结构类别proxy给出三档压力结果与中度情景1亿元调整表。",
                "evidence_artifact_ids": [output_relative],
            },
            "REQ-2020C-06": {
                "claim_id": "CLAIM-2020C-06",
                "claim_text": (
                    "候选输出含未访问test标签声明、primary metric扰动重算与完整约束审计。"
                ),
                "evidence_artifact_ids": [output_relative],
            },
        },
        "robustness_evidence": {
            "failure_cases": [
                "SEVERE_LABEL_IMBALANCE_AND_RATING_CONFOUNDING",
                "UNLABELED_TARGET_DOMAIN_CALIBRATION_UNVERIFIED",
                "INDUSTRY_FIELD_UNAVAILABLE_SHOCK_CATEGORY_PROXY_ONLY",
            ],
            "metric": "validation_composite_loss",
            "metric_direction": "MIN",
            "perturbations": robustness_perturbations,
        },
        "status": "SUCCESS",
        "uncertainty": {
            "domain_shift": shift_summary,
            "finite_labeled_sample": {
                "train_defaults": int(np.sum(y_train)),
                "train_total": len(train_ids),
                "validation_defaults": int(np.sum(y_validation)),
                "validation_total": len(validation_ids),
            },
            "lgd_not_observed": {"tested_values": [0.4, 0.6, 0.8]},
            "shock_severity_not_observed": {
                "scenario_count": 3,
                "scenarios": ["mild", "moderate", "severe"],
            },
        },
        "validation_metrics": validation_metrics,
    }
    if output["final_metrics"][plan["metric"]] != output["validation_metrics"][plan["metric"]]:
        raise ValueError("PRIMARY_METRIC_CONTRACT_MISMATCH")
    output_path = root / args.output
    output_path.write_text(
        json.dumps(output, ensure_ascii=False, indent=2, sort_keys=True, allow_nan=False) + "\n",
        encoding="utf-8",
    )


if __name__ == "__main__":
    main()
