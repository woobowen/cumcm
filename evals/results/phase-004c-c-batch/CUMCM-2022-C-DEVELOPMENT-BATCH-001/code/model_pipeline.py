#!/usr/bin/env python3
"""Answer-sealed, first-party modeling pipeline for one official CUMCM case.

The deterministic case runner supplies ``--case-root``, ``--candidate-id``,
``--seed`` and ``--output``.  This program writes exactly one JSON output and
does not modify raw inputs.  Candidate selection uses only the predeclared
validation score; the unknown samples have no answer labels in this workspace.
"""

from __future__ import annotations

import argparse
import hashlib
import json
import math
import re
from collections import defaultdict
from pathlib import Path
from typing import Any

import numpy as np
from openpyxl import load_workbook
from scipy.stats import chi2_contingency, spearmanr
from sklearn.cluster import AgglomerativeClustering
from sklearn.linear_model import LogisticRegression
from sklearn.metrics import adjusted_rand_score, silhouette_score
from sklearn.neighbors import KNeighborsClassifier
from sklearn.pipeline import make_pipeline
from sklearn.preprocessing import StandardScaler

PIPELINE_VERSION = "2022-c-first-run-v1"
CANDIDATES = (
    "BASELINE_RAW_CENTROID",
    "CLR_RIDGE_WARD",
    "HELLINGER_KNN_COMPLETE",
)
COMPONENTS = (
    "SiO2",
    "Na2O",
    "K2O",
    "CaO",
    "MgO",
    "Al2O3",
    "Fe2O3",
    "CuO",
    "PbO",
    "BaO",
    "P2O5",
    "SrO",
    "SnO2",
    "SO2",
)
TYPE_TO_INT = {"高钾": 0, "铅钡": 1}
INT_TO_TYPE = {value: key for key, value in TYPE_TO_INT.items()}
SPLIT_SALT = "C-TARGET-BATCH-001-POSITION-1-SPLIT-V1"
ZERO_FRACTIONS = (0.25, 0.50, 0.75)


def parse_args() -> argparse.Namespace:
    parser = argparse.ArgumentParser()
    parser.add_argument("--case-root", required=True)
    parser.add_argument("--candidate-id", choices=CANDIDATES, required=True)
    parser.add_argument("--seed", type=int, required=True)
    parser.add_argument("--output", required=True)
    return parser.parse_args()


def safe_path(root: Path, relative: str) -> Path:
    value = Path(relative)
    if value.is_absolute() or ".." in value.parts:
        raise ValueError("CASE_RELATIVE_PATH_REQUIRED")
    resolved = (root / value).resolve()
    resolved.relative_to(root.resolve())
    return resolved


def base_artifact_id(sample_id: str) -> str:
    match = re.match(r"^(\d+)", sample_id)
    if match is None:
        raise ValueError("SAMPLE_ID_WITHOUT_ARTIFACT_PREFIX")
    return match.group(1).zfill(2)


def local_weathering(sample_id: str, surface_status: str) -> str:
    if "未风化点" in sample_id:
        return "无风化"
    if "严重风化点" in sample_id:
        return "严重风化"
    return surface_status


def finite_float(value: Any) -> float:
    result = float(value)
    if not math.isfinite(result):
        raise ValueError("NONFINITE_RESULT")
    return result


def json_safe(value: Any) -> Any:
    if isinstance(value, dict):
        return {str(key): json_safe(item) for key, item in value.items()}
    if isinstance(value, (list, tuple)):
        return [json_safe(item) for item in value]
    if isinstance(value, np.ndarray):
        return json_safe(value.tolist())
    if isinstance(value, (np.integer,)):
        return int(value)
    if isinstance(value, (np.floating, float)):
        return finite_float(value)
    if isinstance(value, (str, int, bool)) or value is None:
        return value
    raise TypeError(f"unsupported JSON value: {type(value).__name__}")


def load_official_workbook(
    case_root: Path,
) -> tuple[list[dict[str, Any]], list[dict[str, Any]], list[dict[str, Any]]]:
    workbook_path = safe_path(case_root, "raw/case_files/附件.xlsx")
    workbook = load_workbook(workbook_path, read_only=True, data_only=True)

    metadata: list[dict[str, Any]] = []
    for row in workbook["表单1"].iter_rows(min_row=2, values_only=True):
        artifact_id = str(row[0]).zfill(2)
        metadata.append(
            {
                "artifact_id": artifact_id,
                "pattern": row[1],
                "glass_type": row[2],
                "color": row[3] if row[3] is not None else "未记录",
                "surface_weathering": row[4],
            }
        )
    by_id = {record["artifact_id"]: record for record in metadata}

    known: list[dict[str, Any]] = []
    for row in workbook["表单2"].iter_rows(min_row=2, values_only=True):
        sample_id = str(row[0])
        artifact_id = base_artifact_id(sample_id)
        values = np.array(
            [0.0 if value is None else float(value) for value in row[1:]], dtype=float
        )
        observed = np.array([value is not None for value in row[1:]], dtype=bool)
        total = float(values.sum())
        meta = by_id[artifact_id]
        known.append(
            {
                **meta,
                "sample_id": sample_id,
                "composition": values,
                "observed": observed,
                "total": total,
                "valid": 85.0 <= total <= 105.0,
                "local_weathering": local_weathering(sample_id, meta["surface_weathering"]),
            }
        )

    unknown: list[dict[str, Any]] = []
    for row in workbook["表单3"].iter_rows(min_row=2, values_only=True):
        values = np.array(
            [0.0 if value is None else float(value) for value in row[2:]], dtype=float
        )
        observed = np.array([value is not None for value in row[2:]], dtype=bool)
        total = float(values.sum())
        unknown.append(
            {
                "sample_id": str(row[0]),
                "surface_weathering": row[1],
                "composition": values,
                "observed": observed,
                "total": total,
                "valid": 85.0 <= total <= 105.0,
            }
        )
    return metadata, known, unknown


def close_composition(values: np.ndarray) -> np.ndarray:
    values = np.asarray(values, dtype=float)
    if np.any(values < 0) or not np.isfinite(values).all() or values.sum() <= 0:
        raise ValueError("INVALID_COMPOSITION")
    return values / values.sum()


def replace_zeros(values: np.ndarray, fraction: float = 0.50) -> np.ndarray:
    closed = close_composition(values)
    positive = closed[closed > 0]
    if positive.size == 0:
        raise ValueError("EMPTY_COMPOSITION")
    replacement = float(positive.min()) * fraction
    replaced = np.where(closed > 0, closed, replacement)
    return close_composition(replaced)


def clr(values: np.ndarray, fraction: float = 0.50) -> np.ndarray:
    logged = np.log(replace_zeros(values, fraction))
    return logged - logged.mean()


def transform(values: np.ndarray, candidate_id: str, fraction: float = 0.50) -> np.ndarray:
    if candidate_id == "BASELINE_RAW_CENTROID":
        return close_composition(values)
    if candidate_id == "CLR_RIDGE_WARD":
        return clr(values, fraction)
    if candidate_id == "HELLINGER_KNN_COMPLETE":
        return np.sqrt(close_composition(values))
    raise ValueError("UNKNOWN_CANDIDATE")


def deterministic_split(metadata: list[dict[str, Any]]) -> dict[str, list[str]]:
    by_type: dict[str, list[str]] = defaultdict(list)
    for record in metadata:
        by_type[record["glass_type"]].append(record["artifact_id"])
    assignment = {"train": [], "validation": [], "test": []}
    for glass_type, identifiers in sorted(by_type.items()):
        ordered = sorted(
            identifiers,
            key=lambda item: hashlib.sha256(
                f"{SPLIT_SALT}:{glass_type}:{item}".encode()
            ).hexdigest(),
        )
        count = len(ordered)
        validation_count = max(2, round(count * 0.20))
        test_count = max(2, round(count * 0.20))
        train_count = count - validation_count - test_count
        if train_count < 2:
            raise ValueError("INSUFFICIENT_STRATIFIED_GROUPS")
        assignment["train"].extend(ordered[:train_count])
        assignment["validation"].extend(ordered[train_count : train_count + validation_count])
        assignment["test"].extend(ordered[train_count + validation_count :])
    return {key: sorted(values) for key, values in assignment.items()}


def artifact_level_dataset(
    metadata: list[dict[str, Any]],
    known: list[dict[str, Any]],
    candidate_id: str,
    fraction: float = 0.50,
) -> tuple[list[str], np.ndarray, np.ndarray]:
    grouped: dict[str, list[np.ndarray]] = defaultdict(list)
    for sample in known:
        if sample["valid"]:
            grouped[sample["artifact_id"]].append(
                transform(sample["composition"], candidate_id, fraction)
            )
    type_by_id = {record["artifact_id"]: TYPE_TO_INT[record["glass_type"]] for record in metadata}
    ids = sorted(grouped)
    matrix = np.vstack([np.median(np.vstack(grouped[item]), axis=0) for item in ids])
    labels = np.array([type_by_id[item] for item in ids], dtype=int)
    return ids, matrix, labels


def fit_type_classifier(candidate_id: str, x: np.ndarray, y: np.ndarray, seed: int) -> Any:
    if candidate_id == "BASELINE_RAW_CENTROID":
        centers = {label: np.mean(x[y == label], axis=0) for label in sorted(set(y.tolist()))}

        class CentroidModel:
            def predict_proba(self, values: np.ndarray) -> np.ndarray:
                distances = np.column_stack(
                    [np.linalg.norm(values - centers[label], axis=1) for label in (0, 1)]
                )
                scale = max(float(np.median(distances)), 1e-9)
                weights = np.exp(-distances / scale)
                return weights / weights.sum(axis=1, keepdims=True)

        return CentroidModel()
    if candidate_id == "CLR_RIDGE_WARD":
        model = make_pipeline(
            StandardScaler(),
            LogisticRegression(C=0.5, solver="liblinear", random_state=seed, max_iter=2000),
        )
        return model.fit(x, y)
    neighbors = max(1, min(5, int(np.sqrt(len(y)))))
    model = KNeighborsClassifier(n_neighbors=neighbors, weights="distance", metric="euclidean")
    return model.fit(x, y)


def brier_loss(y_true: np.ndarray, probability_one: np.ndarray) -> float:
    return finite_float(np.mean((probability_one - y_true.astype(float)) ** 2))


def cramers_v(table: np.ndarray) -> float:
    if table.size == 0 or min(table.shape) < 2 or table.sum() == 0:
        return 0.0
    chi2, _, _, _ = chi2_contingency(table, correction=False)
    denominator = table.sum() * min(table.shape[0] - 1, table.shape[1] - 1)
    return finite_float(math.sqrt(chi2 / denominator)) if denominator > 0 else 0.0


def contingency_result(metadata: list[dict[str, Any]], field: str) -> dict[str, Any]:
    rows = [record for record in metadata if record[field] != "未记录"]
    levels = sorted({record[field] for record in rows})
    weather_levels = ["无风化", "风化"]
    table = np.array(
        [
            [
                sum(
                    record[field] == level and record["surface_weathering"] == weather
                    for record in rows
                )
                for weather in weather_levels
            ]
            for level in levels
        ],
        dtype=int,
    )
    if min(table.shape) >= 2 and np.all(table.sum(axis=0) > 0):
        chi2, p_value, _, expected = chi2_contingency(table, correction=False)
        sparse_expected = int(np.sum(expected < 5))
    else:
        chi2, p_value, sparse_expected = 0.0, 1.0, int(table.size)
    return {
        "field": field,
        "levels": levels,
        "weathering_levels": weather_levels,
        "counts": table.tolist(),
        "n": int(table.sum()),
        "missing_excluded": len(metadata) - len(rows),
        "chi_square": finite_float(chi2),
        "asymptotic_p_value_diagnostic_only": finite_float(p_value),
        "cells_expected_below_5": sparse_expected,
        "cramers_v": cramers_v(table),
        "interpretation_guard": (
            "ASSOCIATION_NOT_CAUSATION;SPARSE_TABLE_REQUIRES_EXACT_OR_PERMUTATION_CHECK"
        ),
    }


def compositional_center(samples: list[np.ndarray], fraction: float = 0.50) -> np.ndarray:
    values = np.vstack([clr(item, fraction) for item in samples])
    center = np.exp(np.median(values, axis=0))
    return close_composition(center)


def weathering_effects(known: list[dict[str, Any]], fraction: float = 0.50) -> dict[str, Any]:
    result: dict[str, Any] = {}
    for glass_type in sorted(TYPE_TO_INT):
        group = [item for item in known if item["valid"] and item["glass_type"] == glass_type]
        unweathered = [
            item["composition"] for item in group if item["local_weathering"] == "无风化"
        ]
        weathered = [item["composition"] for item in group if item["local_weathering"] != "无风化"]
        if not unweathered or not weathered:
            result[glass_type] = {"status": "INSUFFICIENT_GROUPS"}
            continue
        u_center = compositional_center(unweathered, fraction)
        w_center = compositional_center(weathered, fraction)
        delta = clr(w_center, fraction) - clr(u_center, fraction)
        order = np.argsort(np.abs(delta))[::-1]
        result[glass_type] = {
            "unweathered_n": len(unweathered),
            "weathered_n": len(weathered),
            "unweathered_center_percent": {
                name: finite_float(100 * value)
                for name, value in zip(COMPONENTS, u_center, strict=True)
            },
            "weathered_center_percent": {
                name: finite_float(100 * value)
                for name, value in zip(COMPONENTS, w_center, strict=True)
            },
            "largest_absolute_clr_shifts": [
                {
                    "component": COMPONENTS[index],
                    "weathered_minus_unweathered_clr": finite_float(delta[index]),
                }
                for index in order[:6]
            ],
        }
    return result


def restore_weathered_samples(
    known: list[dict[str, Any]], candidate_id: str, fraction: float = 0.50
) -> list[dict[str, Any]]:
    centers: dict[tuple[str, str], np.ndarray] = {}
    for glass_type in sorted(TYPE_TO_INT):
        for status in ("无风化", "风化"):
            members = [
                item["composition"]
                for item in known
                if item["valid"]
                and item["glass_type"] == glass_type
                and ((item["local_weathering"] == "无风化") == (status == "无风化"))
            ]
            if members:
                centers[(glass_type, status)] = compositional_center(members, fraction)
    predictions = []
    for item in known:
        if not item["valid"] or item["local_weathering"] == "无风化":
            continue
        glass_type = item["glass_type"]
        if (glass_type, "无风化") not in centers or (glass_type, "风化") not in centers:
            continue
        if candidate_id == "CLR_RIDGE_WARD":
            shift = clr(centers[(glass_type, "风化")], fraction) - clr(
                centers[(glass_type, "无风化")], fraction
            )
            restored_log = clr(item["composition"], fraction) - shift
            restored = close_composition(np.exp(restored_log))
        elif candidate_id == "HELLINGER_KNN_COMPLETE":
            ratio = np.sqrt(centers[(glass_type, "无风化")]) / np.maximum(
                np.sqrt(centers[(glass_type, "风化")]), 1e-9
            )
            restored = close_composition(
                (np.sqrt(close_composition(item["composition"])) * ratio) ** 2
            )
        else:
            restored = centers[(glass_type, "无风化")]
        predictions.append(
            {
                "sample_id": item["sample_id"],
                "glass_type": glass_type,
                "method_scope": "TYPE_CONDITIONED_BACKCAST_NOT_CAUSAL_IDENTIFICATION",
                "predicted_preweather_percent": {
                    name: finite_float(100 * value)
                    for name, value in zip(COMPONENTS, restored, strict=True)
                },
            }
        )
    return predictions


def cluster_subtypes(
    metadata: list[dict[str, Any]],
    known: list[dict[str, Any]],
    candidate_id: str,
    fraction: float = 0.50,
) -> dict[str, Any]:
    ids, matrix, labels = artifact_level_dataset(metadata, known, candidate_id, fraction)
    output: dict[str, Any] = {}
    for label, glass_type in sorted(INT_TO_TYPE.items()):
        positions = np.where(labels == label)[0]
        local_ids = [ids[index] for index in positions]
        all_values = matrix[positions]
        feature_variance = np.var(all_values, axis=0)
        selected_positions = sorted(
            np.argsort(feature_variance)[::-1][: min(6, all_values.shape[1])].tolist()
        )
        values = all_values[:, selected_positions]
        maximum_k = min(4, len(values) - 1)
        scores: list[dict[str, Any]] = []
        clusterings: dict[int, np.ndarray] = {}
        linkage = "ward" if candidate_id != "HELLINGER_KNN_COMPLETE" else "complete"
        for k in range(2, maximum_k + 1):
            fitted = AgglomerativeClustering(n_clusters=k, linkage=linkage).fit_predict(values)
            if len(set(fitted.tolist())) < 2:
                score = -1.0
            else:
                score = finite_float(silhouette_score(values, fitted, metric="euclidean"))
            clusterings[k] = fitted
            scores.append({"k": k, "silhouette": score})
        chosen_k = max(scores, key=lambda item: (item["silhouette"], -item["k"]))["k"]
        chosen = clusterings[chosen_k]
        assignments = {
            item: f"{glass_type}-S{int(group) + 1}"
            for item, group in zip(local_ids, chosen, strict=True)
        }
        output[glass_type] = {
            "candidate_k_diagnostics": scores,
            "selected_k": chosen_k,
            "selection_rule": "MAX_SILHOUETTE_THEN_SMALLEST_K;K_IN_2_TO_4",
            "linkage": linkage,
            "artifact_assignments": assignments,
            "feature_space": candidate_id,
            "selected_components": [COMPONENTS[index] for index in selected_positions],
            "feature_selection_rule": "TOP_SIX_WITHIN_TYPE_VARIANCES_BEFORE_CLUSTERING",
        }
    return output


def cluster_sensitivity(
    metadata: list[dict[str, Any]], known: list[dict[str, Any]], candidate_id: str
) -> dict[str, Any]:
    reference = cluster_subtypes(metadata, known, candidate_id, 0.50)
    comparisons: dict[str, list[float]] = {glass_type: [] for glass_type in TYPE_TO_INT}
    for fraction in ZERO_FRACTIONS:
        alternative = cluster_subtypes(metadata, known, candidate_id, fraction)
        for glass_type in comparisons:
            ref_assign = reference[glass_type]["artifact_assignments"]
            alt_assign = alternative[glass_type]["artifact_assignments"]
            common = sorted(set(ref_assign) & set(alt_assign))
            ref_labels = [ref_assign[item] for item in common]
            alt_labels = [alt_assign[item] for item in common]
            comparisons[glass_type].append(
                finite_float(adjusted_rand_score(ref_labels, alt_labels))
            )
    return {
        glass_type: {
            "zero_replacement_fractions": list(ZERO_FRACTIONS),
            "adjusted_rand_indices": values,
            "minimum_adjusted_rand_index": min(values),
        }
        for glass_type, values in comparisons.items()
    }


def predict_unknown(
    metadata: list[dict[str, Any]],
    known: list[dict[str, Any]],
    unknown: list[dict[str, Any]],
    candidate_id: str,
    seed: int,
) -> list[dict[str, Any]]:
    ids, matrix, labels = artifact_level_dataset(metadata, known, candidate_id)
    split = deterministic_split(metadata)
    fitting_ids = set(split["train"] + split["validation"])
    positions = [index for index, item in enumerate(ids) if item in fitting_ids]
    model = fit_type_classifier(candidate_id, matrix[positions], labels[positions], seed)
    rng = np.random.default_rng(seed)
    output = []
    for item in unknown:
        base = transform(item["composition"], candidate_id).reshape(1, -1)
        probability = float(model.predict_proba(base)[0, 1])
        perturbed_probabilities: list[float] = []
        for _ in range(200):
            noise = rng.normal(0.0, 0.01, size=len(COMPONENTS))
            perturbed = close_composition(
                np.maximum(close_composition(item["composition"]) + noise, 1e-8)
            )
            features = transform(perturbed, candidate_id).reshape(1, -1)
            perturbed_probabilities.append(float(model.predict_proba(features)[0, 1]))
        predicted = INT_TO_TYPE[int(probability >= 0.5)]
        stable_fraction = float(
            np.mean([(value >= 0.5) == (probability >= 0.5) for value in perturbed_probabilities])
        )
        output.append(
            {
                "sample_id": item["sample_id"],
                "predicted_type": predicted,
                "probability_lead_barium": finite_float(probability),
                "decision_margin": finite_float(abs(probability - 0.5)),
                "perturbation_same_class_fraction": finite_float(stable_fraction),
                "perturbation_probability_interval": [
                    finite_float(np.quantile(perturbed_probabilities, 0.025)),
                    finite_float(np.quantile(perturbed_probabilities, 0.975)),
                ],
            }
        )
    return output


def association_matrices(known: list[dict[str, Any]], fraction: float = 0.50) -> dict[str, Any]:
    matrices: dict[str, np.ndarray] = {}
    output: dict[str, Any] = {}
    for glass_type in sorted(TYPE_TO_INT):
        values = np.vstack(
            [
                clr(item["composition"], fraction)
                for item in known
                if item["valid"] and item["glass_type"] == glass_type
            ]
        )
        correlation, _ = spearmanr(values, axis=0)
        correlation = np.asarray(correlation, dtype=float)
        correlation = np.nan_to_num(correlation, nan=0.0, posinf=0.0, neginf=0.0)
        np.fill_diagonal(correlation, 1.0)
        matrices[glass_type] = correlation
        pairs = []
        for left in range(len(COMPONENTS)):
            for right in range(left + 1, len(COMPONENTS)):
                pairs.append(
                    {
                        "left": COMPONENTS[left],
                        "right": COMPONENTS[right],
                        "spearman_clr": finite_float(correlation[left, right]),
                    }
                )
        pairs.sort(key=lambda item: (-abs(item["spearman_clr"]), item["left"], item["right"]))
        output[glass_type] = {"sample_n": len(values), "strongest_absolute_pairs": pairs[:12]}
    first, second = sorted(TYPE_TO_INT)
    difference = matrices[first] - matrices[second]
    differences = []
    for left in range(len(COMPONENTS)):
        for right in range(left + 1, len(COMPONENTS)):
            differences.append(
                {
                    "left": COMPONENTS[left],
                    "right": COMPONENTS[right],
                    "spearman_difference": finite_float(difference[left, right]),
                }
            )
    differences.sort(
        key=lambda item: (-abs(item["spearman_difference"]), item["left"], item["right"])
    )
    output["between_type_difference"] = {
        "frobenius_norm": finite_float(np.linalg.norm(difference, ord="fro")),
        "largest_absolute_pair_differences": differences[:12],
        "guard": "COMPOSITIONAL_ASSOCIATION_IS_DESCRIPTIVE;NO_CAUSAL_DIRECTION",
    }
    return output


def validation_score(
    metadata: list[dict[str, Any]], known: list[dict[str, Any]], candidate_id: str, seed: int
) -> tuple[float, dict[str, Any]]:
    ids, matrix, labels = artifact_level_dataset(metadata, known, candidate_id)
    split = deterministic_split(metadata)
    train_ids = set(split["train"])
    validation_ids = set(split["validation"])
    train_positions = [index for index, item in enumerate(ids) if item in train_ids]
    validation_positions = [index for index, item in enumerate(ids) if item in validation_ids]
    if not train_positions or not validation_positions:
        raise ValueError("EMPTY_EFFECTIVE_SPLIT")
    model = fit_type_classifier(
        candidate_id, matrix[train_positions], labels[train_positions], seed
    )
    probabilities = model.predict_proba(matrix[validation_positions])[:, 1]
    loss = brier_loss(labels[validation_positions], probabilities)
    predictions = (probabilities >= 0.5).astype(int)
    accuracy = finite_float(np.mean(predictions == labels[validation_positions]))
    return loss, {
        "group_unit": "ARTIFACT_ID",
        "split_assignment": split,
        "effective_train_ids": [ids[index] for index in train_positions],
        "effective_validation_ids": [ids[index] for index in validation_positions],
        "invalid_composition_artifacts_absent": sorted(
            set(split["train"] + split["validation"]) - set(ids)
        ),
        "validation_brier_loss": loss,
        "validation_accuracy_diagnostic": accuracy,
        "test_labels_accessed": False,
    }


def main() -> int:
    args = parse_args()
    case_root = Path(args.case_root).resolve()
    output_path = safe_path(case_root, args.output)
    metadata, known, unknown = load_official_workbook(case_root)
    loss, validation = validation_score(metadata, known, args.candidate_id, args.seed)
    invalid_known = [item["sample_id"] for item in known if not item["valid"]]

    payload = {
        "schema_version": "1.0.0",
        "pipeline_version": PIPELINE_VERSION,
        "candidate_id": args.candidate_id,
        "seed": args.seed,
        "status": "SUCCESS",
        "validation_metrics": {"validation_composite_loss": loss},
        "validation_detail": validation,
        "data_scope": {
            "known_rows": len(known),
            "valid_known_rows": sum(item["valid"] for item in known),
            "invalid_known_sample_ids": invalid_known,
            "unknown_rows": len(unknown),
            "unknown_valid_rows": sum(item["valid"] for item in unknown),
            "blank_cell_policy": "NONDETECT_ZERO_WITH_MULTIPLICATIVE_REPLACEMENT_SENSITIVITY",
            "artifact_grouping": True,
        },
        "question_1": {
            "weathering_associations": [
                contingency_result(metadata, "glass_type"),
                contingency_result(metadata, "pattern"),
                contingency_result(metadata, "color"),
            ],
            "type_conditioned_composition_effects": weathering_effects(known),
            "preweathering_backcasts": restore_weathered_samples(known, args.candidate_id),
        },
        "question_2": {
            "type_classifier": {
                "candidate_id": args.candidate_id,
                "validation": validation,
                "decision_threshold": 0.5,
            },
            "within_type_subtypes": cluster_subtypes(metadata, known, args.candidate_id),
            "subtype_sensitivity": cluster_sensitivity(metadata, known, args.candidate_id),
        },
        "question_3": {
            "unknown_type_predictions": predict_unknown(
                metadata, known, unknown, args.candidate_id, args.seed
            ),
            "sensitivity_design": "200 SEEDED ONE_PERCENT_SIMPLEX_PERTURBATIONS_PER_UNKNOWN",
            "answer_labels_available": False,
        },
        "question_4": association_matrices(known),
        "limitations": [
            "Historical model-prior exposure is unverifiable.",
            "Blank cells are nondetections, not ordinary missing-at-random values.",
            (
                "Only two within-artifact weathered versus unweathered pairs are apparent; "
                "backcasts rely on type-level exchangeability."
            ),
            (
                "Multiple sampling points from one artifact are not independent and are grouped "
                "for validation."
            ),
            "Sparse contingency tables make asymptotic p-values diagnostic only.",
            "Subtype count is exploratory and must pass sensitivity review before final selection.",
            (
                "Association differences are descriptive and do not establish causal chemical "
                "mechanisms."
            ),
        ],
    }
    output_path.parent.mkdir(parents=True, exist_ok=True)
    output_path.write_text(
        json.dumps(
            json_safe(payload), ensure_ascii=False, indent=2, sort_keys=True, allow_nan=False
        )
        + "\n",
        encoding="utf-8",
    )
    return 0


if __name__ == "__main__":
    raise SystemExit(main())
