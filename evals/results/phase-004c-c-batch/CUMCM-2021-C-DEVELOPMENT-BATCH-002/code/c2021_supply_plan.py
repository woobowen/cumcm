#!/usr/bin/env python3
"""Answer-sealed, first-party end-to-end model for CUMCM 2021 problem C.

The program uses only the official case inputs.  Candidate generation and scoring use
the preregistered train/validation periods; W217-W240 are never inspected here.
"""

from __future__ import annotations

import argparse
import json
from dataclasses import dataclass
from pathlib import Path
from typing import Any

import numpy as np
from c2021_feasibility import CONSUMPTION, PURCHASE_PRICE, verify_plan
from openpyxl import load_workbook
from scipy.optimize import Bounds, LinearConstraint, linprog, milp

CANDIDATES = {
    "BASELINE_MEAN_GREEDY": {
        "capacity_quantile": 0.50,
        "ratio_method": "MEAN",
        "loss_method": "MEAN",
        "allocation": "GREEDY",
    },
    "ROBUST_QUANTILE_LEXICOGRAPHIC": {
        "capacity_quantile": 0.75,
        "ratio_method": "Q20",
        "loss_method": "Q75",
        "allocation": "LINEAR_PROGRAM",
    },
    "SCENARIO_CVAR_PORTFOLIO": {
        "capacity_quantile": 0.75,
        "ratio_method": "BOOTSTRAP_LOWER_TAIL",
        "loss_method": "UPPER_TAIL_MEAN",
        "allocation": "LINEAR_PROGRAM",
    },
}
SUPPLIER_FILE = "raw/case_files/附件1 近5年402家供应商的相关数据.xlsx"
CARRIER_FILE = "raw/case_files/附件2 近5年8家转运商的相关数据.xlsx"
TRAIN = slice(0, 168)
VALIDATION = slice(168, 216)
WEEKLY_DEMAND = 28200.0
CARRIER_CAPACITY = 6000.0
EPS = 1e-8


@dataclass(frozen=True)
class CaseData:
    supplier_ids: list[str]
    supplier_types: list[str]
    order: np.ndarray
    supply: np.ndarray
    carrier_ids: list[str]
    loss_pct: np.ndarray


def _sheet_matrix(
    path: Path, sheet: str, first_numeric_column: int
) -> tuple[list[str], list[str], np.ndarray]:
    workbook = load_workbook(path, read_only=True, data_only=True)
    worksheet = workbook[sheet]
    rows = [row for row in worksheet.iter_rows(values_only=True) if row[0] is not None]
    headers = [str(value) for value in rows[0]]
    identifiers = [str(row[0]) for row in rows[1:]]
    matrix = np.asarray(
        [[float(value) for value in row[first_numeric_column:]] for row in rows[1:]],
        dtype=float,
    )
    workbook.close()
    return headers, identifiers, matrix


def load_case_data(case_root: Path) -> CaseData:
    supplier_path = case_root / SUPPLIER_FILE
    carrier_path = case_root / CARRIER_FILE
    order_headers, order_ids, order = _sheet_matrix(supplier_path, "企业的订货量（m³）", 2)
    supply_headers, supply_ids, supply = _sheet_matrix(supplier_path, "供应商的供货量（m³）", 2)
    workbook = load_workbook(supplier_path, read_only=True, data_only=True)
    worksheet = workbook["企业的订货量（m³）"]
    supplier_types = [
        str(row[1]) for row in list(worksheet.iter_rows(values_only=True))[1:] if row[0] is not None
    ]
    workbook.close()
    carrier_headers, carrier_ids, loss_pct = _sheet_matrix(carrier_path, "运输损耗率（%）", 1)
    expected_weeks = [f"W{week:03d}" for week in range(1, 241)]
    if (
        order_ids != supply_ids
        or order_headers[2:] != expected_weeks
        or supply_headers[2:] != expected_weeks
        or carrier_headers[1:] != expected_weeks
        or order.shape != (402, 240)
        or supply.shape != (402, 240)
        or loss_pct.shape != (8, 240)
        or any(material not in CONSUMPTION for material in supplier_types)
        or np.any(~np.isfinite(order))
        or np.any(~np.isfinite(supply))
        or np.any(~np.isfinite(loss_pct))
        or np.any(order < 0)
        or np.any(supply < 0)
        or np.any(loss_pct < 0)
    ):
        raise ValueError("OFFICIAL_INPUT_SCHEMA_INVALID")
    return CaseData(order_ids, supplier_types, order, supply, carrier_ids, loss_pct)


def _safe_quantile(values: np.ndarray, quantile: float, default: float = 0.0) -> float:
    values = values[np.isfinite(values)]
    return default if values.size == 0 else float(np.quantile(values, quantile))


def estimate_suppliers(data: CaseData, candidate_id: str, seed: int) -> dict[str, np.ndarray]:
    config = CANDIDATES[candidate_id]
    rng = np.random.default_rng(seed)
    order = data.order[:, TRAIN]
    supply = data.supply[:, TRAIN]
    count = len(data.supplier_ids)
    order_capacity = np.zeros(count)
    ratio = np.zeros(count)
    reliability = np.zeros(count)
    stability = np.zeros(count)
    activity = np.zeros(count)
    mean_product = np.zeros(count)
    for index in range(count):
        ordered = order[index] > 0
        positive_orders = order[index, ordered]
        ratios = np.clip(supply[index, ordered] / positive_orders, 0.0, 2.0)
        order_capacity[index] = _safe_quantile(positive_orders, float(config["capacity_quantile"]))
        if ratios.size:
            if config["ratio_method"] == "MEAN":
                ratio[index] = float(np.mean(ratios))
            elif config["ratio_method"] == "Q20":
                ratio[index] = _safe_quantile(ratios, 0.20)
            else:
                boot = np.asarray(
                    [np.mean(rng.choice(ratios, size=ratios.size, replace=True)) for _ in range(64)]
                )
                cutoff = np.quantile(boot, 0.20)
                ratio[index] = float(np.mean(boot[boot <= cutoff]))
            reliability[index] = float(np.mean(ratios >= 0.95))
        positives = supply[index, supply[index] > 0]
        activity[index] = float(np.mean(supply[index] > 0))
        if positives.size:
            stability[index] = 1.0 / (
                1.0 + float(np.std(positives)) / max(float(np.mean(positives)), EPS)
            )
        mean_product[index] = (
            float(np.mean(supply[index])) / CONSUMPTION[data.supplier_types[index]]
        )
    ratio = np.clip(ratio, 0.0, 1.5)
    delivered_capacity = order_capacity * ratio
    return {
        "order_capacity": order_capacity,
        "delivery_ratio": ratio,
        "delivered_capacity": delivered_capacity,
        "reliability": reliability,
        "stability": stability,
        "activity": activity,
        "mean_product": mean_product,
    }


def estimate_carrier_losses(data: CaseData, candidate_id: str) -> np.ndarray:
    method = CANDIDATES[candidate_id]["loss_method"]
    losses = np.zeros(len(data.carrier_ids))
    for index, row in enumerate(data.loss_pct[:, TRAIN]):
        positive = row[row > 0] / 100.0
        if positive.size == 0:
            raise ValueError(f"CARRIER_POSITIVE_LOSS_MISSING:{data.carrier_ids[index]}")
        if method == "MEAN":
            losses[index] = float(np.mean(positive))
        elif method == "Q75":
            losses[index] = _safe_quantile(positive, 0.75)
        else:
            cutoff = np.quantile(positive, 0.80)
            losses[index] = float(np.mean(positive[positive >= cutoff]))
    return losses


def importance_scores(
    data: CaseData, estimates: dict[str, np.ndarray]
) -> tuple[np.ndarray, list[dict[str, Any]]]:
    feature_names = ["mean_product", "reliability", "stability", "activity"]
    weights = np.asarray([0.45, 0.25, 0.15, 0.15])
    normalized: list[np.ndarray] = []
    for name in feature_names:
        values = estimates[name]
        span = float(np.max(values) - np.min(values))
        normalized.append(
            np.zeros_like(values) if span <= EPS else (values - np.min(values)) / span
        )
    score = np.vstack(normalized).T @ weights
    ranked = sorted(
        range(len(data.supplier_ids)), key=lambda i: (-float(score[i]), data.supplier_ids[i])
    )
    top50 = [
        {
            "rank": rank,
            "supplier_id": data.supplier_ids[index],
            "material_type": data.supplier_types[index],
            "importance_score": round(float(score[index]), 9),
        }
        for rank, index in enumerate(ranked[:50], 1)
    ]
    return score, top50


def _effective_coefficients(data: CaseData, losses: np.ndarray) -> np.ndarray:
    consumption = np.asarray([CONSUMPTION[value] for value in data.supplier_types])
    return (1.0 - losses[None, :]) / consumption[:, None]


def minimum_supplier_set(
    data: CaseData,
    estimates: dict[str, np.ndarray],
    losses: np.ndarray,
    importance: np.ndarray,
    candidate_id: str,
) -> tuple[np.ndarray, str]:
    capacity = estimates["delivered_capacity"]
    effective_best = (
        capacity
        * float(np.max(1.0 - losses))
        / np.asarray([CONSUMPTION[value] for value in data.supplier_types])
    )
    if candidate_id == "BASELINE_MEAN_GREEDY":
        order = sorted(
            range(len(capacity)),
            key=lambda index: (-float(effective_best[index]), data.supplier_ids[index]),
        )
        selected: list[int] = []
        total = 0.0
        for index in order:
            if effective_best[index] <= EPS:
                continue
            selected.append(index)
            total += float(effective_best[index])
            if total + EPS >= WEEKLY_DEMAND:
                break
        if total + EPS < WEEKLY_DEMAND:
            raise ValueError("BASELINE_MINIMUM_SUPPLIER_INFEASIBLE")
        return np.asarray(selected, dtype=int), "GREEDY_UPPER_BOUND"

    supplier_count = len(data.supplier_ids)
    carrier_count = len(data.carrier_ids)
    variable_count = supplier_count + supplier_count * carrier_count
    objective = np.zeros(variable_count)
    objective[:supplier_count] = 1.0 + 1e-6 * (1.0 - importance)
    objective[supplier_count:] = 1e-10
    lower = np.zeros(variable_count)
    upper = np.full(variable_count, np.inf)
    upper[:supplier_count] = 1.0
    rows: list[np.ndarray] = []
    lb: list[float] = []
    ub: list[float] = []
    for supplier in range(supplier_count):
        row = np.zeros(variable_count)
        row[supplier] = -capacity[supplier]
        start = supplier_count + supplier * carrier_count
        row[start : start + carrier_count] = 1.0
        rows.append(row)
        lb.append(-np.inf)
        ub.append(0.0)
    for carrier in range(carrier_count):
        row = np.zeros(variable_count)
        row[supplier_count + carrier :: carrier_count] = 1.0
        rows.append(row)
        lb.append(-np.inf)
        ub.append(CARRIER_CAPACITY)
    demand_row = np.zeros(variable_count)
    demand_row[supplier_count:] = -_effective_coefficients(data, losses).ravel()
    rows.append(demand_row)
    lb.append(-np.inf)
    ub.append(-WEEKLY_DEMAND)
    result = milp(
        objective,
        integrality=np.r_[np.ones(supplier_count), np.zeros(supplier_count * carrier_count)],
        bounds=Bounds(lower, upper),
        constraints=LinearConstraint(np.vstack(rows), np.asarray(lb), np.asarray(ub)),
        options={"time_limit": 120.0, "mip_rel_gap": 0.0},
    )
    if not result.success or result.x is None:
        raise ValueError(f"MINIMUM_SUPPLIER_MILP_FAILED:{result.status}")
    return np.flatnonzero(result.x[:supplier_count] > 0.5), "MILP_EXACT_CARDINALITY"


def allocate_lp(
    data: CaseData,
    estimates: dict[str, np.ndarray],
    losses: np.ndarray,
    *,
    objective_mode: str,
    demand: float | None,
    allowed_suppliers: np.ndarray | None = None,
) -> np.ndarray:
    supplier_count = len(data.supplier_ids)
    carrier_count = len(data.carrier_ids)
    capacity = estimates["delivered_capacity"].copy()
    if allowed_suppliers is not None:
        mask = np.zeros(supplier_count, dtype=bool)
        mask[allowed_suppliers] = True
        capacity[~mask] = 0.0
    effective = _effective_coefficients(data, losses)
    ratio = np.maximum(estimates["delivery_ratio"], EPS)
    prices = np.asarray([PURCHASE_PRICE[value] for value in data.supplier_types])
    if objective_mode == "CAPACITY_MAX":
        objective = -effective
    else:
        procurement = prices[:, None] / ratio[:, None]
        raw_and_loss = 0.05 + 0.50 * losses[None, :]
        if objective_mode == "A_PRIORITY":
            material_penalty = np.asarray(
                [
                    0.0 if value == "A" else 2.0 if value == "B" else 20.0
                    for value in data.supplier_types
                ]
            )[:, None]
        else:
            material_penalty = np.zeros((supplier_count, 1))
        objective = procurement + raw_and_loss + material_penalty
    a_ub: list[np.ndarray] = []
    b_ub: list[float] = []
    for supplier in range(supplier_count):
        row = np.zeros((supplier_count, carrier_count))
        row[supplier, :] = 1.0
        a_ub.append(row.ravel())
        b_ub.append(float(capacity[supplier]))
    for carrier in range(carrier_count):
        row = np.zeros((supplier_count, carrier_count))
        row[:, carrier] = 1.0
        a_ub.append(row.ravel())
        b_ub.append(CARRIER_CAPACITY)
    if demand is not None:
        a_ub.append(-effective.ravel())
        b_ub.append(-float(demand))
    result = linprog(
        objective.ravel(),
        A_ub=np.vstack(a_ub),
        b_ub=np.asarray(b_ub),
        bounds=(0.0, None),
        method="highs",
    )
    if not result.success or result.x is None:
        raise ValueError(f"ALLOCATION_LP_FAILED:{objective_mode}:{result.status}")
    return result.x.reshape(supplier_count, carrier_count)


def allocate_greedy(
    data: CaseData,
    estimates: dict[str, np.ndarray],
    losses: np.ndarray,
    *,
    objective_mode: str,
    demand: float | None,
    allowed_suppliers: np.ndarray | None = None,
) -> np.ndarray:
    supplier_count = len(data.supplier_ids)
    carrier_count = len(data.carrier_ids)
    allocation = np.zeros((supplier_count, carrier_count))
    remaining_carrier = np.full(carrier_count, CARRIER_CAPACITY)
    allowed = (
        set(range(supplier_count))
        if allowed_suppliers is None
        else set(map(int, allowed_suppliers))
    )
    effective = _effective_coefficients(data, losses)
    ratio = np.maximum(estimates["delivery_ratio"], EPS)
    prices = np.asarray([PURCHASE_PRICE[value] for value in data.supplier_types])
    if objective_mode == "CAPACITY_MAX":
        supplier_order = sorted(
            allowed,
            key=lambda index: (-float(np.max(effective[index])), data.supplier_ids[index]),
        )
    elif objective_mode == "A_PRIORITY":
        material_rank = {"A": 0, "B": 1, "C": 2}
        supplier_order = sorted(
            allowed,
            key=lambda index: (
                material_rank[data.supplier_types[index]],
                prices[index] / ratio[index],
                data.supplier_ids[index],
            ),
        )
    else:
        supplier_order = sorted(
            allowed,
            key=lambda index: (prices[index] / ratio[index], data.supplier_ids[index]),
        )
    achieved = 0.0
    for supplier in supplier_order:
        remaining_supplier = float(estimates["delivered_capacity"][supplier])
        for carrier in sorted(range(carrier_count), key=lambda j: (losses[j], data.carrier_ids[j])):
            if remaining_supplier <= EPS or remaining_carrier[carrier] <= EPS:
                continue
            volume = min(remaining_supplier, remaining_carrier[carrier])
            if demand is not None:
                needed = max(0.0, demand - achieved)
                volume = min(volume, needed / effective[supplier, carrier])
            if volume <= EPS:
                continue
            allocation[supplier, carrier] += volume
            remaining_supplier -= volume
            remaining_carrier[carrier] -= volume
            achieved += volume * effective[supplier, carrier]
            if demand is not None and achieved + 1e-6 >= demand:
                return allocation
    if demand is not None and achieved + 1e-6 < demand:
        raise ValueError("GREEDY_ALLOCATION_INFEASIBLE")
    return allocation


def sparse_plan(
    data: CaseData,
    estimates: dict[str, np.ndarray],
    allocation: np.ndarray,
) -> dict[str, Any]:
    supplied = np.sum(allocation, axis=1)
    ratio = estimates["delivery_ratio"]
    orders = [
        {"supplier_id": data.supplier_ids[i], "volume_m3": round(float(supplied[i] / ratio[i]), 6)}
        for i in range(len(data.supplier_ids))
        if supplied[i] > EPS and ratio[i] > EPS
    ]
    transport = [
        {
            "supplier_id": data.supplier_ids[i],
            "carrier_id": data.carrier_ids[j],
            "volume_m3": round(float(allocation[i, j]), 6),
        }
        for i in range(len(data.supplier_ids))
        for j in range(len(data.carrier_ids))
        if allocation[i, j] > EPS
    ]
    return {"orders": orders, "transport": transport, "repeat_for_weeks": list(range(1, 25))}


def validation_score(
    data: CaseData,
    plan: dict[str, Any],
    estimates: dict[str, np.ndarray],
    forecast_losses: np.ndarray,
) -> dict[str, float]:
    supplier_index = {value: index for index, value in enumerate(data.supplier_ids)}
    carrier_index = {value: index for index, value in enumerate(data.carrier_ids)}
    order_map = {record["supplier_id"]: float(record["volume_m3"]) for record in plan["orders"]}
    transport_share: dict[tuple[str, str], float] = {}
    shipped_total: dict[str, float] = {}
    for record in plan["transport"]:
        supplier = str(record["supplier_id"])
        shipped_total[supplier] = shipped_total.get(supplier, 0.0) + float(record["volume_m3"])
    for record in plan["transport"]:
        key = (str(record["supplier_id"]), str(record["carrier_id"]))
        transport_share[key] = float(record["volume_m3"]) / shipped_total[key[0]]
    weekly_shortage: list[float] = []
    weekly_capacity_excess: list[float] = []
    for week in range(VALIDATION.start, VALIDATION.stop):
        carrier_load = np.zeros(len(data.carrier_ids))
        effective_received = 0.0
        for supplier, ordered_volume in order_map.items():
            i = supplier_index[supplier]
            historical_order = data.order[i, week]
            actual_ratio = (
                float(np.clip(data.supply[i, week] / historical_order, 0.0, 2.0))
                if historical_order > 0
                else float(estimates["delivery_ratio"][i])
            )
            actual_supply = ordered_volume * actual_ratio
            for carrier in data.carrier_ids:
                share = transport_share.get((supplier, carrier), 0.0)
                if share <= 0.0:
                    continue
                j = carrier_index[carrier]
                shipped = actual_supply * share
                carrier_load[j] += shipped
                observed_loss = data.loss_pct[j, week] / 100.0
                loss = float(observed_loss if observed_loss > 0 else forecast_losses[j])
                effective_received += shipped * (1.0 - loss) / CONSUMPTION[data.supplier_types[i]]
        weekly_shortage.append(max(0.0, WEEKLY_DEMAND - effective_received) / WEEKLY_DEMAND)
        weekly_capacity_excess.append(
            float(np.maximum(carrier_load - CARRIER_CAPACITY, 0.0).sum())
            / (CARRIER_CAPACITY * len(data.carrier_ids))
        )
    purchase_cost = sum(
        volume * PURCHASE_PRICE[data.supplier_types[supplier_index[supplier]]]
        for supplier, volume in order_map.items()
    )
    normalized_cost = purchase_cost / WEEKLY_DEMAND
    score = (
        normalized_cost
        + 100.0 * float(np.mean(weekly_shortage))
        + 100.0 * float(np.mean(weekly_capacity_excess))
    )
    return {
        "validation_penalized_cost_per_effective_m3": round(score, 9),
        "normalized_purchase_cost": round(normalized_cost, 9),
        "mean_shortage_fraction": round(float(np.mean(weekly_shortage)), 9),
        "maximum_shortage_fraction": round(float(np.max(weekly_shortage)), 9),
        "mean_carrier_excess_fraction": round(float(np.mean(weekly_capacity_excess)), 9),
    }


def solve(case_root: Path, candidate_id: str, seed: int) -> dict[str, Any]:
    if candidate_id not in CANDIDATES:
        raise ValueError("CANDIDATE_ID_UNKNOWN")
    data = load_case_data(case_root)
    estimates = estimate_suppliers(data, candidate_id, seed)
    losses = estimate_carrier_losses(data, candidate_id)
    importance, top50 = importance_scores(data, estimates)
    selected, cardinality_status = minimum_supplier_set(
        data, estimates, losses, importance, candidate_id
    )
    allocator = (
        allocate_greedy if CANDIDATES[candidate_id]["allocation"] == "GREEDY" else allocate_lp
    )
    q2_allocation = allocator(
        data,
        estimates,
        losses,
        objective_mode="ECONOMIC",
        demand=WEEKLY_DEMAND,
        allowed_suppliers=selected,
    )
    q3_allocation = allocator(
        data, estimates, losses, objective_mode="A_PRIORITY", demand=WEEKLY_DEMAND
    )
    q4_allocation = allocator(data, estimates, losses, objective_mode="CAPACITY_MAX", demand=None)
    plans = {
        "question_2": sparse_plan(data, estimates, q2_allocation),
        "question_3": sparse_plan(data, estimates, q3_allocation),
        "question_4": sparse_plan(data, estimates, q4_allocation),
    }
    type_map = dict(zip(data.supplier_ids, data.supplier_types, strict=True))
    ratio_map = dict(zip(data.supplier_ids, map(float, estimates["delivery_ratio"]), strict=True))
    loss_map = dict(zip(data.carrier_ids, map(float, losses), strict=True))
    feasibility = {
        name: verify_plan(
            plan,
            supplier_types=type_map,
            delivery_ratios=ratio_map,
            carrier_loss_rates=loss_map,
            weekly_demand_product_m3=(None if name == "question_4" else WEEKLY_DEMAND),
            carrier_capacity_m3=CARRIER_CAPACITY,
        )
        for name, plan in plans.items()
    }
    if not all(record["feasible"] for record in feasibility.values()):
        raise ValueError("INDEPENDENT_FEASIBILITY_RECALCULATION_FAILED")
    validation = validation_score(data, plans["question_2"], estimates, losses)
    return {
        "candidate_id": candidate_id,
        "status": "SUCCESS",
        "seed": seed,
        "answer_access_status": "SEALED",
        "test_accessed": False,
        "split_usage": {
            "train": "W001-W168",
            "validation": "W169-W216",
            "test": "W217-W240_UNACCESSED",
        },
        "question_1": {
            "importance_model": "MINMAX_WEIGHTED_CAPABILITY_RELIABILITY_STABILITY_ACTIVITY",
            "top_50_suppliers": top50,
        },
        "question_2": {
            "minimum_supplier_count": int(len(selected)),
            "minimum_supplier_ids": [data.supplier_ids[i] for i in selected],
            "cardinality_status": cardinality_status,
            "weekly_plan_repeated_for_24_weeks": plans["question_2"],
        },
        "question_3": {"weekly_plan_repeated_for_24_weeks": plans["question_3"]},
        "question_4": {
            "weekly_capacity_product_m3": feasibility["question_4"][
                "weekly_effective_received_product_m3"
            ],
            "weekly_capacity_increase_product_m3": round(
                feasibility["question_4"]["weekly_effective_received_product_m3"] - WEEKLY_DEMAND, 6
            ),
            "weekly_plan_repeated_for_24_weeks": plans["question_4"],
        },
        "independent_feasibility": feasibility,
        "validation": validation,
        "metric_name": "validation_penalized_cost_per_effective_m3",
        "metric_value": validation["validation_penalized_cost_per_effective_m3"],
        "limitations": [
            "Future supplier behavior is assumed stable relative to answer-sealed "
            "historical observations.",
            "Zero carrier loss observations denote no transport and are excluded "
            "from loss estimation.",
            "Initial effective inventory is assumed to equal two weeks of target production.",
            "The one-supplier/one-carrier rule is treated as a soft preference and "
            "split counts are reported.",
            "Normalized costs are used because absolute transport and storage "
            "tariffs are not supplied.",
        ],
    }


def main() -> int:
    parser = argparse.ArgumentParser()
    parser.add_argument("--case-root", required=True)
    parser.add_argument("--candidate-id", required=True, choices=sorted(CANDIDATES))
    parser.add_argument("--seed", required=True, type=int)
    parser.add_argument("--output", required=True)
    args = parser.parse_args()
    case_root = Path(args.case_root).resolve()
    output_path = (case_root / args.output).resolve()
    output_path.relative_to(case_root)
    result = solve(case_root, args.candidate_id, args.seed)
    output_path.parent.mkdir(parents=True, exist_ok=True)
    output_path.write_text(
        json.dumps(result, ensure_ascii=False, indent=2, sort_keys=True, allow_nan=False) + "\n",
        encoding="utf-8",
    )
    return 0


if __name__ == "__main__":
    raise SystemExit(main())
