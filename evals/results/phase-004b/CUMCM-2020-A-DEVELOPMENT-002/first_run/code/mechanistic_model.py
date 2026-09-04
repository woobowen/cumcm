#!/usr/bin/env python3
"""First-party mechanistic furnace model for the sealed Development run."""

from __future__ import annotations

import argparse
import json
import math
from pathlib import Path
from typing import Any

import numpy as np
from openpyxl import load_workbook
from scipy.linalg import expm
from scipy.optimize import least_squares

AMBIENT_C = 25.0
ZONE_LENGTH_CM = 30.5
GAP_LENGTH_CM = 5.0
ENTRY_EXIT_LENGTH_CM = 25.0
TOTAL_LENGTH_CM = 2 * ENTRY_EXIT_LENGTH_CM + 11 * ZONE_LENGTH_CM + 10 * GAP_LENGTH_CM
PROCESS_LIMITS = {
    "max_rise_c_per_s": (0.0, 3.0),
    "min_fall_c_per_s": (-3.0, 0.0),
    "rising_150_190_s": (60.0, 120.0),
    "above_217_s": (40.0, 90.0),
    "peak_c": (240.0, 250.0),
}
EXPERIMENT_SETTINGS = [175.0] * 5 + [195.0, 235.0] + [255.0] * 2 + [25.0] * 2
Q1_SETTINGS = [173.0] * 5 + [198.0, 230.0] + [257.0] * 2 + [25.0] * 2
Q2_SETTINGS = [182.0] * 5 + [203.0, 237.0] + [254.0] * 2 + [25.0] * 2
Q1_SPEED_CM_PER_MIN = 78.0
EXPERIMENT_SPEED_CM_PER_MIN = 70.0
MODEL_IDS = {
    "BASELINE_FIRST_ORDER": "single-capacitance first-order thermal lag",
    "PRIMARY_TWO_NODE": "two-capacitance surface-core thermal network",
    "CONTROL_ASYMMETRIC_FIRST_ORDER": "first-order lag with separate heating/cooling constants",
}
REQUIREMENT_CLAIMS = {
    "REQ-2020A-Q1": (
        "The selected run contains a simulated center-temperature curve, four requested checkpoint "
        "temperatures, and 0.5-second figure-ready values for the specified operating settings."
    ),
    "REQ-2020A-Q2": (
        "The selected run reports the largest speed found feasible under every registered process "
        "limit for the fixed Q2 zone settings."
    ),
    "REQ-2020A-Q3": (
        "The selected run reports a feasible bounded direct-search solution minimizing registered "
        "area above 217 degrees Celsius, with settings, speed, curve and residuals."
    ),
    "REQ-2020A-Q4": (
        "The selected run reports a feasible area-bounded solution minimizing the registered "
        "peak-centered symmetry metric, with settings, speed, curve and residuals."
    ),
}


def load_measurements(case_root: Path) -> tuple[np.ndarray, np.ndarray]:
    workbook = load_workbook(
        case_root / "data/raw/case_files/附件.xlsx", read_only=True, data_only=True
    )
    sheet = workbook.active
    rows = [row for row in sheet.iter_rows(min_row=2, values_only=True) if row[0] is not None]
    time_s = np.asarray([float(row[0]) for row in rows], dtype=float)
    temperature_c = np.asarray([float(row[1]) for row in rows], dtype=float)
    if len(time_s) != 709 or not np.all(np.diff(time_s) > 0):
        raise ValueError("MEASUREMENT_SCHEMA_INVALID")
    return time_s, temperature_c


def zone_centers_cm() -> np.ndarray:
    return np.asarray(
        [
            ENTRY_EXIT_LENGTH_CM + index * (ZONE_LENGTH_CM + GAP_LENGTH_CM) + ZONE_LENGTH_CM / 2
            for index in range(11)
        ],
        dtype=float,
    )


def environment_temperature(position_cm: np.ndarray, settings_c: list[float]) -> np.ndarray:
    anchors_x = np.concatenate(([0.0], zone_centers_cm(), [TOTAL_LENGTH_CM]))
    anchors_t = np.asarray([AMBIENT_C, *settings_c, AMBIENT_C], dtype=float)
    return np.interp(position_cm, anchors_x, anchors_t)


def _two_node_coefficients(parameters: np.ndarray, dt_s: float) -> tuple[np.ndarray, np.ndarray]:
    air_rate, surface_core_rate, core_rate = parameters
    matrix = np.asarray(
        [
            [-(air_rate + surface_core_rate), surface_core_rate],
            [core_rate, -core_rate],
        ],
        dtype=float,
    )
    forcing = np.asarray([air_rate, 0.0], dtype=float)
    transition = expm(matrix * dt_s)
    gain = np.linalg.solve(matrix, (transition - np.eye(2)) @ forcing)
    return transition, gain


def simulate(
    candidate_id: str,
    parameters: np.ndarray,
    settings_c: list[float],
    speed_cm_per_min: float,
    *,
    dt_s: float = 0.25,
) -> tuple[np.ndarray, np.ndarray, np.ndarray]:
    speed_cm_per_s = speed_cm_per_min / 60.0
    duration_s = TOTAL_LENGTH_CM / speed_cm_per_s
    time_s = np.arange(0.0, duration_s + 0.5 * dt_s, dt_s)
    position_cm = np.minimum(speed_cm_per_s * time_s, TOTAL_LENGTH_CM)
    oven_c = environment_temperature(position_cm, settings_c)
    if candidate_id == "PRIMARY_TWO_NODE":
        state = np.asarray([AMBIENT_C, AMBIENT_C], dtype=float)
        board_c = np.empty_like(time_s)
        board_c[0] = state[1]
        transition, gain = _two_node_coefficients(parameters, dt_s)
        for index in range(1, len(time_s)):
            forcing = 0.5 * (oven_c[index - 1] + oven_c[index])
            state = transition @ state + gain * forcing
            board_c[index] = state[1]
    else:
        board_c = np.empty_like(time_s)
        board_c[0] = AMBIENT_C
        for index in range(1, len(time_s)):
            forcing = 0.5 * (oven_c[index - 1] + oven_c[index])
            if candidate_id == "BASELINE_FIRST_ORDER":
                tau_s = parameters[0]
            else:
                tau_s = parameters[0] if forcing >= board_c[index - 1] else parameters[1]
            decay = math.exp(-dt_s / tau_s)
            board_c[index] = forcing + (board_c[index - 1] - forcing) * decay
    return time_s, board_c, oven_c


def calibrated_parameters(
    candidate_id: str, measured_time_s: np.ndarray, measured_temperature_c: np.ndarray
) -> tuple[np.ndarray, dict[str, Any]]:
    train_end = int(0.80 * len(measured_time_s))
    validation_end = int(0.90 * len(measured_time_s))
    train_time = measured_time_s[:train_end]
    train_temperature = measured_temperature_c[:train_end]

    if candidate_id == "BASELINE_FIRST_ORDER":
        initial = np.log([35.0])
        lower = np.log([3.0])
        upper = np.log([150.0])
    elif candidate_id == "CONTROL_ASYMMETRIC_FIRST_ORDER":
        initial = np.log([35.0, 25.0])
        lower = np.log([3.0, 3.0])
        upper = np.log([150.0, 150.0])
    else:
        initial = np.log([0.07, 0.035, 0.025])
        lower = np.log([0.002, 0.001, 0.001])
        upper = np.log([0.5, 0.5, 0.5])

    def residual(log_parameters: np.ndarray) -> np.ndarray:
        parameters = np.exp(log_parameters)
        time_s, board_c, _ = simulate(
            candidate_id, parameters, EXPERIMENT_SETTINGS, EXPERIMENT_SPEED_CM_PER_MIN
        )
        predicted = np.interp(train_time, time_s, board_c)
        return predicted - train_temperature

    fitted = least_squares(
        residual,
        initial,
        bounds=(lower, upper),
        max_nfev=160,
        ftol=1e-10,
        xtol=1e-10,
        gtol=1e-10,
    )
    parameters = np.exp(fitted.x)
    time_s, board_c, _ = simulate(
        candidate_id, parameters, EXPERIMENT_SETTINGS, EXPERIMENT_SPEED_CM_PER_MIN
    )
    prediction = np.interp(measured_time_s, time_s, board_c)
    validation_slice = slice(train_end, validation_end)
    validation_rmse = float(
        np.sqrt(
            np.mean(
                np.square(prediction[validation_slice] - measured_temperature_c[validation_slice])
            )
        )
    )
    train_rmse = float(np.sqrt(np.mean(np.square(prediction[:train_end] - train_temperature))))
    return parameters, {
        "calibration_method": "bounded nonlinear least squares on the first 80% of ordered samples",
        "cost": float(fitted.cost),
        "function_evaluations": int(fitted.nfev),
        "gradient_optimality": float(fitted.optimality),
        "parameters": [float(value) for value in parameters],
        "solver_message": str(fitted.message),
        "solver_success": bool(fitted.success),
        "train_end_index_exclusive": train_end,
        "train_rmse_c": train_rmse,
        "validation_end_index_exclusive": validation_end,
        "validation_rmse_c": validation_rmse,
    }


def process_metrics(time_s: np.ndarray, temperature_c: np.ndarray) -> dict[str, float]:
    dt = np.diff(time_s)
    slopes = np.diff(temperature_c) / dt
    midpoint_temperature = 0.5 * (temperature_c[:-1] + temperature_c[1:])
    rising_150_190 = float(
        np.sum(
            dt[(midpoint_temperature >= 150.0) & (midpoint_temperature <= 190.0) & (slopes >= 0)]
        )
    )
    above_217 = float(np.sum(dt[midpoint_temperature > 217.0]))
    excess = np.maximum(temperature_c - 217.0, 0.0)
    area = float(np.trapezoid(excess, time_s))
    return {
        "above_217_s": above_217,
        "area_above_217_c_s": area,
        "max_rise_c_per_s": float(np.max(slopes)),
        "min_fall_c_per_s": float(np.min(slopes)),
        "peak_c": float(np.max(temperature_c)),
        "rising_150_190_s": rising_150_190,
    }


def constraint_violations(metrics: dict[str, float]) -> dict[str, float]:
    values: dict[str, float] = {}
    for name, (minimum, maximum) in PROCESS_LIMITS.items():
        value = metrics[name]
        values[f"{name}_below"] = max(0.0, minimum - value)
        values[f"{name}_above"] = max(0.0, value - maximum)
    return values


def is_feasible(violations: dict[str, float], tolerance: float = 1e-6) -> bool:
    return max(violations.values(), default=0.0) <= tolerance


def symmetry_metric(time_s: np.ndarray, temperature_c: np.ndarray) -> float:
    peak_index = int(np.argmax(temperature_c))
    high_indices = np.flatnonzero(temperature_c >= 217.0)
    if len(high_indices) < 3 or peak_index <= high_indices[0] or peak_index >= high_indices[-1]:
        return float("inf")
    half_width = min(
        time_s[peak_index] - time_s[high_indices[0]], time_s[high_indices[-1]] - time_s[peak_index]
    )
    offsets = np.linspace(0.0, half_width, 121)
    left = np.interp(time_s[peak_index] - offsets, time_s, temperature_c)
    right = np.interp(time_s[peak_index] + offsets, time_s, temperature_c)
    return float(np.sqrt(np.mean(np.square(left - right))))


def settings_from_vector(vector: np.ndarray) -> tuple[list[float], float]:
    group_1_5, zone_6, zone_7, group_8_9, speed = [float(value) for value in vector]
    return [group_1_5] * 5 + [zone_6, zone_7] + [group_8_9] * 2 + [25.0] * 2, speed


def evaluate_vector(
    candidate_id: str, parameters: np.ndarray, vector: np.ndarray
) -> tuple[dict[str, float], dict[str, float], float]:
    settings, speed = settings_from_vector(vector)
    time_s, temperature_c, _ = simulate(candidate_id, parameters, settings, speed)
    metrics = process_metrics(time_s, temperature_c)
    return metrics, constraint_violations(metrics), symmetry_metric(time_s, temperature_c)


def coordinate_refine(
    candidate_id: str,
    parameters: np.ndarray,
    initial: np.ndarray,
    objective: str,
    area_limit: float | None = None,
) -> tuple[np.ndarray, dict[str, float], dict[str, float], float, int]:
    lower = np.asarray([165.0, 185.0, 225.0, 245.0, 65.0])
    upper = np.asarray([185.0, 205.0, 245.0, 265.0, 100.0])
    current = initial.copy()
    metrics, violations, symmetry = evaluate_vector(candidate_id, parameters, current)
    step = np.asarray([2.0, 2.0, 2.0, 2.0, 2.0])
    evaluations = 1

    def score(candidate_metrics: dict[str, float], candidate_symmetry: float) -> float:
        if objective == "area":
            return candidate_metrics["area_above_217_c_s"]
        return candidate_symmetry

    for _ in range(6):
        improved = True
        while improved:
            improved = False
            for index in range(5):
                for direction in (-1.0, 1.0):
                    proposal = current.copy()
                    proposal[index] = np.clip(
                        proposal[index] + direction * step[index], lower[index], upper[index]
                    )
                    proposal_metrics, proposal_violations, proposal_symmetry = evaluate_vector(
                        candidate_id, parameters, proposal
                    )
                    evaluations += 1
                    area_ok = (
                        area_limit is None or proposal_metrics["area_above_217_c_s"] <= area_limit
                    )
                    if (
                        is_feasible(proposal_violations)
                        and area_ok
                        and score(proposal_metrics, proposal_symmetry) + 1e-9
                        < score(metrics, symmetry)
                    ):
                        current = proposal
                        metrics = proposal_metrics
                        violations = proposal_violations
                        symmetry = proposal_symmetry
                        improved = True
        step *= 0.5
    return current, metrics, violations, symmetry, evaluations


def constrained_search(
    candidate_id: str, parameters: np.ndarray, seed: int
) -> tuple[dict[str, Any], dict[str, Any]]:
    rng = np.random.default_rng(seed)
    lower = np.asarray([165.0, 185.0, 225.0, 245.0, 65.0])
    upper = np.asarray([185.0, 205.0, 245.0, 265.0, 100.0])
    samples = lower + rng.random((1400, 5)) * (upper - lower)
    samples = np.vstack(
        [
            samples,
            np.asarray([175.0, 195.0, 235.0, 255.0, 70.0]),
            np.asarray([182.0, 203.0, 237.0, 254.0, 78.0]),
        ]
    )
    feasible: list[tuple[np.ndarray, dict[str, float], dict[str, float], float]] = []
    for vector in samples:
        metrics, violations, symmetry = evaluate_vector(candidate_id, parameters, vector)
        if is_feasible(violations) and math.isfinite(symmetry):
            feasible.append((vector.copy(), metrics, violations, symmetry))
    if not feasible:
        return (
            {
                "status": "INFEASIBLE",
                "sample_count": int(len(samples)),
                "failure_reason": "NO_FEASIBLE_SAMPLE_FOUND",
            },
            {
                "status": "INFEASIBLE",
                "sample_count": int(len(samples)),
                "failure_reason": "NO_FEASIBLE_SAMPLE_FOUND",
            },
        )

    q3_start = min(feasible, key=lambda item: (item[1]["area_above_217_c_s"], tuple(item[0])))[0]
    q3_vector, q3_metrics, q3_violations, q3_symmetry, q3_local = coordinate_refine(
        candidate_id, parameters, q3_start, "area"
    )
    area_limit = 1.05 * q3_metrics["area_above_217_c_s"]
    symmetry_pool = [item for item in feasible if item[1]["area_above_217_c_s"] <= area_limit]
    q4_start = min(
        symmetry_pool, key=lambda item: (item[3], item[1]["area_above_217_c_s"], tuple(item[0]))
    )[0]
    q4_vector, q4_metrics, q4_violations, q4_symmetry, q4_local = coordinate_refine(
        candidate_id, parameters, q4_start, "symmetry", area_limit
    )

    def result(
        vector: np.ndarray,
        metrics: dict[str, float],
        violations: dict[str, float],
        symmetry: float,
        local_evaluations: int,
    ) -> dict[str, Any]:
        settings, speed = settings_from_vector(vector)
        time_s, temperature_c, _ = simulate(candidate_id, parameters, settings, speed, dt_s=0.5)
        return {
            "constraint_violations": rounded_mapping(violations),
            "curve_0_5s": curve_records(time_s, temperature_c),
            "feasible": is_feasible(violations),
            "metrics": rounded_mapping(metrics),
            "search": {
                "algorithm": "seeded bounded random search plus coordinate refinement",
                "global_optimum_proven": False,
                "local_evaluations": local_evaluations,
                "sample_count": int(len(samples)),
                "seed": seed,
            },
            "settings_c": [round(value, 6) for value in settings],
            "speed_cm_per_min": round(speed, 6),
            "symmetry_rmse_c": round(symmetry, 8),
        }

    return (
        result(q3_vector, q3_metrics, q3_violations, q3_symmetry, q3_local),
        result(q4_vector, q4_metrics, q4_violations, q4_symmetry, q4_local),
    )


def maximum_feasible_speed(candidate_id: str, parameters: np.ndarray) -> dict[str, Any]:
    feasible_speeds: list[tuple[float, dict[str, float], dict[str, float]]] = []
    for speed in np.linspace(65.0, 100.0, 141):
        time_s, temperature_c, _ = simulate(candidate_id, parameters, Q2_SETTINGS, float(speed))
        metrics = process_metrics(time_s, temperature_c)
        violations = constraint_violations(metrics)
        if is_feasible(violations):
            feasible_speeds.append((float(speed), metrics, violations))
    if not feasible_speeds:
        return {"status": "INFEASIBLE", "grid_step_cm_per_min": 0.25}
    speed, metrics, violations = max(feasible_speeds, key=lambda item: item[0])
    lower = speed
    upper = min(100.0, speed + 0.25)
    for _ in range(18):
        midpoint = 0.5 * (lower + upper)
        time_s, temperature_c, _ = simulate(candidate_id, parameters, Q2_SETTINGS, midpoint)
        midpoint_metrics = process_metrics(time_s, temperature_c)
        midpoint_violations = constraint_violations(midpoint_metrics)
        if is_feasible(midpoint_violations):
            lower, metrics, violations = midpoint, midpoint_metrics, midpoint_violations
        else:
            upper = midpoint
    return {
        "constraint_violations": rounded_mapping(violations),
        "feasible": True,
        "grid_step_cm_per_min": 0.25,
        "metrics": rounded_mapping(metrics),
        "speed_cm_per_min": round(lower, 8),
        "status": "SUCCESS",
    }


def curve_records(time_s: np.ndarray, temperature_c: np.ndarray) -> list[dict[str, float]]:
    return [
        {"temperature_c": round(float(temperature), 6), "time_s": round(float(time), 6)}
        for time, temperature in zip(time_s, temperature_c, strict=True)
    ]


def rounded_mapping(values: dict[str, float]) -> dict[str, float]:
    return {key: round(float(value), 8) for key, value in values.items()}


def q1_result(candidate_id: str, parameters: np.ndarray) -> dict[str, Any]:
    time_s, temperature_c, _ = simulate(
        candidate_id, parameters, Q1_SETTINGS, Q1_SPEED_CM_PER_MIN, dt_s=0.25
    )
    checkpoint_positions = {
        "zone_3_midpoint": ENTRY_EXIT_LENGTH_CM
        + 2 * (ZONE_LENGTH_CM + GAP_LENGTH_CM)
        + ZONE_LENGTH_CM / 2,
        "zone_6_midpoint": ENTRY_EXIT_LENGTH_CM
        + 5 * (ZONE_LENGTH_CM + GAP_LENGTH_CM)
        + ZONE_LENGTH_CM / 2,
        "zone_7_midpoint": ENTRY_EXIT_LENGTH_CM
        + 6 * (ZONE_LENGTH_CM + GAP_LENGTH_CM)
        + ZONE_LENGTH_CM / 2,
        "zone_8_end": ENTRY_EXIT_LENGTH_CM + 7 * (ZONE_LENGTH_CM + GAP_LENGTH_CM) + ZONE_LENGTH_CM,
    }
    speed_cm_per_s = Q1_SPEED_CM_PER_MIN / 60.0
    checkpoints = {
        name: {
            "position_cm": round(position, 6),
            "temperature_c": round(
                float(np.interp(position / speed_cm_per_s, time_s, temperature_c)), 6
            ),
            "time_s": round(position / speed_cm_per_s, 6),
        }
        for name, position in checkpoint_positions.items()
    }
    sample_time = np.arange(0.0, time_s[-1] + 1e-9, 0.5)
    sample_temperature = np.interp(sample_time, time_s, temperature_c)
    return {
        "checkpoints": checkpoints,
        "curve_0_5s": curve_records(sample_time, sample_temperature),
        "metrics": rounded_mapping(process_metrics(time_s, temperature_c)),
        "settings_c": Q1_SETTINGS,
        "speed_cm_per_min": Q1_SPEED_CM_PER_MIN,
    }


def numerical_validation(
    candidate_id: str, parameters: np.ndarray, q3: dict[str, Any]
) -> dict[str, Any]:
    settings = q3.get("settings_c", Q1_SETTINGS)
    speed = float(q3.get("speed_cm_per_min", Q1_SPEED_CM_PER_MIN))
    coarse_t, coarse_y, _ = simulate(candidate_id, parameters, settings, speed, dt_s=0.5)
    fine_t, fine_y, _ = simulate(candidate_id, parameters, settings, speed, dt_s=0.125)
    common = np.arange(0.0, min(coarse_t[-1], fine_t[-1]) + 1e-9, 0.5)
    difference = np.interp(common, coarse_t, coarse_y) - np.interp(common, fine_t, fine_y)
    return {
        "coarse_dt_s": 0.5,
        "fine_dt_s": 0.125,
        "max_abs_temperature_difference_c": round(float(np.max(np.abs(difference))), 8),
        "rmse_temperature_difference_c": round(float(np.sqrt(np.mean(np.square(difference)))), 8),
        "unit_conversion_checked": "speed cm/min converted once to cm/s",
    }


def main() -> int:
    parser = argparse.ArgumentParser()
    parser.add_argument("--case-root", required=True)
    parser.add_argument("--candidate-id", choices=sorted(MODEL_IDS), required=True)
    parser.add_argument("--seed", type=int, required=True)
    parser.add_argument("--output", required=True)
    args = parser.parse_args()
    case_root = Path(args.case_root)
    measured_time_s, measured_temperature_c = load_measurements(case_root)
    parameters, calibration = calibrated_parameters(
        args.candidate_id, measured_time_s, measured_temperature_c
    )
    q1 = q1_result(args.candidate_id, parameters)
    q2 = maximum_feasible_speed(args.candidate_id, parameters)
    q3, q4 = constrained_search(args.candidate_id, parameters, args.seed)
    run_success = (
        calibration["solver_success"]
        and q2.get("status") == "SUCCESS"
        and q3.get("feasible") is True
        and q4.get("feasible") is True
    )
    output = {
        "candidate_id": args.candidate_id,
        "actual_implemented_method": MODEL_IDS[args.candidate_id],
        "calibration": calibration,
        "constraint_summary": {
            "q2": q2.get("constraint_violations", {}),
            "q3": q3.get("constraint_violations", {}),
            "q4": q4.get("constraint_violations", {}),
        },
        "declared_method": MODEL_IDS[args.candidate_id],
        "failure_conditions": [
            "no feasible process window found inside registered bounds",
            "calibration solver failure or active parameter bound",
            "time-step sensitivity above the registered tolerance",
            "unmodeled board-specific heat transfer invalidates fitted parameters",
        ],
        "model_id": args.candidate_id,
        "numerical_validation": numerical_validation(args.candidate_id, parameters, q3),
        "process_limits": PROCESS_LIMITS,
        "q1": q1,
        "q2": q2,
        "q3": q3,
        "q4": q4,
        "requirement_claims": REQUIREMENT_CLAIMS,
        "seed": args.seed,
        "solver_status": "SUCCESS" if run_success else "FAILED_HARD_GATE",
        "status": "SUCCESS" if run_success else "PARTIAL",
        "validation_metrics": {"rmse_c": round(calibration["validation_rmse_c"], 8)},
    }
    output_path = case_root / args.output
    output_path.parent.mkdir(parents=True, exist_ok=True)
    output_path.write_text(
        json.dumps(output, ensure_ascii=False, sort_keys=True, indent=2, allow_nan=False) + "\n",
        encoding="utf-8",
    )
    print(
        json.dumps(
            {
                "candidate_id": args.candidate_id,
                "q2_status": q2.get("status"),
                "q3_status": q3.get("status", "SUCCESS"),
                "q4_status": q4.get("status", "SUCCESS"),
                "run_success": run_success,
                "validation_rmse_c": output["validation_metrics"]["rmse_c"],
            },
            sort_keys=True,
        )
    )
    return 0 if run_success else 2


if __name__ == "__main__":
    raise SystemExit(main())
